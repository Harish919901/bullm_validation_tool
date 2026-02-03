"""
Validation API Router
Handles file upload, validation, export, and save operations
"""

import os
import uuid
import csv
import io
from typing import Dict, Any
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from models.schemas import (
    ValidationResult,
    ValidationResponse,
    RuleInfo,
    RulesResponse
)
from validators.quote_win import ExcelValidator as QuoteWinValidator
from validators.bom_matrix import ExcelValidator as BOMValidator

router = APIRouter(prefix="/api", tags=["validation"])

# In-memory session storage (for simplicity)
# In production, use Redis or database
sessions: Dict[str, Any] = {}


@router.post("/validate", response_model=ValidationResponse)
async def validate_file(
    file: UploadFile = File(...),
    validator_type: str = Form(default="QUOTE_WIN")
):
    """
    Upload and validate an Excel file

    Args:
        file: Excel file to validate
        validator_type: Type of validation - "BOM" or "QUOTE_WIN"

    Returns:
        ValidationResponse with results
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    # Generate session ID
    session_id = str(uuid.uuid4())

    # Save uploaded file temporarily
    temp_dir = os.path.join(os.path.dirname(__file__), "..", "temp")
    os.makedirs(temp_dir, exist_ok=True)
    temp_path = os.path.join(temp_dir, f"{session_id}_{file.filename}")

    try:
        # Write file to temp location
        content = await file.read()
        with open(temp_path, "wb") as f:
            f.write(content)

        # Run validation based on type
        results = []

        if validator_type == "QUOTE_WIN":
            results = run_quote_win_validation(temp_path)
        elif validator_type == "BOM":
            # BOM Matrix validation placeholder
            results = run_bom_validation(temp_path)
        else:
            raise HTTPException(status_code=400, detail=f"Invalid validator type: {validator_type}")

        # Calculate stats
        passed = sum(1 for r in results if r.status == "PASS")
        failed = len(results) - passed

        # Store session data
        sessions[session_id] = {
            "file_path": temp_path,
            "file_name": file.filename,
            "validator_type": validator_type,
            "results": results
        }

        return ValidationResponse(
            session_id=session_id,
            results=results,
            passed=passed,
            failed=failed,
            total=len(results),
            file_name=file.filename,
            validator_type=validator_type
        )

    except Exception as e:
        # Clean up temp file on error
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise HTTPException(status_code=500, detail=str(e))


def run_quote_win_validation(file_path: str) -> list[ValidationResult]:
    """Run Quote Win validation and return results"""
    validator = QuoteWinValidator(file_path)

    if not validator.load_file():
        raise HTTPException(status_code=500, detail="Failed to load Excel file")

    # Run all validations
    validator.validate_headers()
    validator.validate_project_name()
    validator.validate_filters()
    validator.validate_awards()

    # Convert to ValidationResult objects
    results = []
    validation_results = validator.validation_results

    # Header Validation
    header_val = validation_results['header_validation']
    results.append(ValidationResult(
        rule_name="Rule 1: Header Validation",
        sheet_name=validator.input_sheet.title,
        status="FAIL" if not header_val['passed'] else "PASS",
        expected="All required headers present in rows 12 and 16",
        actual=header_val['message'] if not header_val['passed'] else "All headers present",
        location=f"Row {validator.summary_header_row} and Row {validator.header_row}",
        validator_type="QUOTE_WIN"
    ))

    # Project Name Validation
    project_val = validation_results['project_name_validation']
    results.append(ValidationResult(
        rule_name="Rule 2: Project Name Validation",
        sheet_name=validator.input_sheet.title,
        status="FAIL" if not project_val['passed'] else "PASS",
        expected="Project names match between header and data section",
        actual=project_val['message'] if not project_val['passed'] else "Project names match",
        location=f"Row {validator.project_row}, Column {validator.project_value_col}",
        validator_type="QUOTE_WIN"
    ))

    # Filter Validation
    filter_val = validation_results['filter_validation']
    results.append(ValidationResult(
        rule_name="Rule 3: Filter Validation",
        sheet_name=validator.input_sheet.title,
        status="FAIL" if not filter_val['passed'] else "PASS",
        expected="No filters applied",
        actual=filter_val['message'] if not filter_val['passed'] else "No filters applied",
        location="Cell A1",
        validator_type="QUOTE_WIN"
    ))

    # Award Validation
    award_val = validation_results['award_validation']
    if not award_val['passed']:
        actual_parts = []
        award_column_results = award_val.get('award_column_results', {})

        for award_num in sorted(award_column_results.keys(), key=lambda x: int(x)):
            col_result = award_column_results[award_num]
            col_issues = []

            no_award_parts = [p for p in col_result.get('no_award_parts', []) if p and p.strip()]
            multiple_award_parts = [p for p in col_result.get('multiple_award_parts', []) if p and p.strip()]

            if no_award_parts:
                col_issues.append(f"Missing award for Part Numbers: {', '.join(no_award_parts)}")
            if multiple_award_parts:
                col_issues.append(f"Multiple awards for Part Numbers: {', '.join(multiple_award_parts)}")

            if col_issues:
                actual_parts.append(f"Award #{award_num}: {'; '.join(col_issues)}")

        actual_message = " | ".join(actual_parts) if actual_parts else award_val['message']

        results.append(ValidationResult(
            rule_name="Rule 4: Award Validation",
            sheet_name=validator.input_sheet.title,
            status="FAIL",
            expected="Each Award column must have Award value of 100 for each part number",
            actual=actual_message,
            location=f"Data rows starting at Row {validator.data_start_row}",
            validator_type="QUOTE_WIN"
        ))
    else:
        results.append(ValidationResult(
            rule_name="Rule 4: Award Validation",
            sheet_name=validator.input_sheet.title,
            status="PASS",
            expected="Each Award column must have Award value of 100 for each part number",
            actual="All Award columns validated - each part number has exactly one award per column",
            location=f"Data rows starting at Row {validator.data_start_row}",
            validator_type="QUOTE_WIN"
        ))

    return results


def run_bom_validation(file_path: str) -> list[ValidationResult]:
    """Run BOM Matrix validation with all 19 rules"""
    validator = BOMValidator(file_path)
    bom_results = validator.run_all_validations()

    # Convert to ValidationResult objects
    results = []
    for r in bom_results:
        results.append(ValidationResult(
            rule_name=r.rule_name,
            sheet_name=r.sheet_name,
            status=r.status,
            expected=r.expected,
            actual=r.actual,
            location=r.location,
            validator_type="BOM"
        ))

    return results


@router.get("/rules/{validator_type}", response_model=RulesResponse)
async def get_rules(validator_type: str):
    """Get rules information for a validator type"""

    if validator_type == "QUOTE_WIN":
        rules = [
            RuleInfo(rule_num="Rule 1", title="Header Validation",
                    description="Validates all required headers are present in summary row (Row 12) and main header row (Row 16) with correct patterns"),
            RuleInfo(rule_num="Rule 2", title="Project Name Validation",
                    description="Validates that the project name matches between the project row and the data section"),
            RuleInfo(rule_num="Rule 3", title="Filter Validation",
                    description="Checks if any filters are applied to the Excel file and flags if found"),
            RuleInfo(rule_num="Rule 4", title="Award Validation",
                    description="Validates that each unique part number has at least one Award column with value 100"),
        ]
    elif validator_type == "BOM":
        rules = [
            RuleInfo(rule_num="Rule 1", title="Header Validation",
                    description="Validates 'Ext Part Vol Price (Splits) #{X} (Conv.)' headers in CBOM VL sheets"),
            RuleInfo(rule_num="Rule 2", title="Quoted MPN Validation",
                    description="Validates 'Quoted MPN' is present under 'Corrected MPN Mentioned' section"),
            # Add more BOM rules as needed
        ]
    else:
        raise HTTPException(status_code=400, detail=f"Invalid validator type: {validator_type}")

    return RulesResponse(validator_type=validator_type, rules=rules)


@router.post("/export/csv")
async def export_csv(session_id: str = Form(...)):
    """Export validation results to CSV"""

    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]
    results = session["results"]

    # Create CSV in memory
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Rule Name', 'Sheet Name', 'Status', 'Expected', 'Actual', 'Location'])

    for r in results:
        writer.writerow([r.rule_name, r.sheet_name, r.status, r.expected, r.actual, r.location])

    output.seek(0)

    return StreamingResponse(
        io.BytesIO(output.getvalue().encode('utf-8')),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=validation_report.csv"}
    )


@router.post("/export/excel")
async def export_excel(session_id: str = Form(...)):
    """Export validation results to Excel"""

    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]
    results = session["results"]

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    # Styles
    header_fill = PatternFill(start_color="4A7C59", end_color="4A7C59", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    pass_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    fail_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")

    # Headers
    headers = ['Rule Name', 'Sheet Name', 'Status', 'Expected', 'Actual', 'Location']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row, r in enumerate(results, 2):
        ws.cell(row=row, column=1, value=r.rule_name)
        ws.cell(row=row, column=2, value=r.sheet_name)
        status_cell = ws.cell(row=row, column=3, value=r.status)
        status_cell.fill = pass_fill if r.status == "PASS" else fail_fill
        ws.cell(row=row, column=4, value=r.expected)
        ws.cell(row=row, column=5, value=r.actual)
        ws.cell(row=row, column=6, value=r.location)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 15

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=validation_report.xlsx"}
    )


@router.post("/save")
async def save_validated_file(session_id: str = Form(...)):
    """Save validated file with comments (Quote Win only)"""

    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Session not found")

    session = sessions[session_id]

    if session["validator_type"] != "QUOTE_WIN":
        raise HTTPException(status_code=400, detail="Save is only available for Quote Win validator")

    file_path = session["file_path"]
    file_name = session["file_name"]

    # Read the validated workbook
    try:
        with open(file_path, "rb") as f:
            content = f.read()

        output_name = file_name.replace('.xlsx', '_validated.xlsx')

        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={output_name}"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/session/{session_id}")
async def delete_session(session_id: str):
    """Clean up session and temp files"""

    if session_id in sessions:
        session = sessions[session_id]
        file_path = session.get("file_path")

        if file_path and os.path.exists(file_path):
            os.remove(file_path)

        del sessions[session_id]

    return {"message": "Session deleted"}

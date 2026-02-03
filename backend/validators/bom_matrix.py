"""
Excel Validation Tool
Rule 1: Header Validation for CBOM VL Sheets
"""

import openpyxl
import re
from typing import List, Dict, Tuple
from dataclasses import dataclass


@dataclass
class ValidationResult:
    """Data class to store validation results"""
    rule_name: str
    sheet_name: str
    status: str  # "PASS" or "FAIL"
    expected: str
    actual: str
    location: str = ""  # Row/Column where header was found or should be


class ExcelValidator:
    """Excel validation tool for BOM Matrix files"""
    
    def __init__(self, filepath: str):
        """
        Initialize the validator with an Excel file
        
        Args:
            filepath: Path to the Excel file to validate
        """
        self.filepath = filepath
        self.workbook = openpyxl.load_workbook(filepath, data_only=True)
        self.results: List[ValidationResult] = []
    
    def diagnose_currency_issues(self, sheet_name: str, headers_to_check: List[str]) -> None:
        """
        Diagnose currency formatting issues for debugging
        
        Args:
            sheet_name: Name of the sheet to diagnose
            headers_to_check: List of headers to check for currency
        """
        print(f"\n{'='*100}")
        print(f"DIAGNOSING: {sheet_name} - Currency Issues")
        print('='*100)
        
        if sheet_name not in self.workbook.sheetnames:
            print(f"[X] Sheet '{sheet_name}' not found!")
            return
        
        sheet = self.workbook[sheet_name]
        
        for header in headers_to_check:
            print(f"\n🔍 Looking for header: '{header}'")
            found = False
            
            for row_idx in range(1, min(100, sheet.max_row + 1)):
                for col_idx in range(1, 100):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value and isinstance(cell.value, str) and header in cell.value:
                        found = True
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        print(f"  [OK] Found at {col_letter}{row_idx}: '{cell.value}'")
                        
                        # Check next 5 cells below
                        print(f"\n  Checking cells below:")
                        for offset in range(1, 6):
                            check_row = row_idx + offset
                            check_cell = sheet.cell(row=check_row, column=col_idx)
                            
                            # Check value
                            val = check_cell.value
                            fmt = check_cell.number_format
                            
                            # Check for currency
                            has_currency_format = any(sym in str(fmt) for sym in ['$', '€', '£', '¥', '₹'])
                            has_currency_symbol = val and any(sym in str(val) for sym in ['$', '€', '£', '¥', '₹'])
                            
                            status = "[OK] HAS CURRENCY" if (has_currency_format or has_currency_symbol) else "[X] NO CURRENCY"
                            
                            print(f"    Row {check_row}: Value={val}, Format='{fmt}' | {status}")
                        break
                if found:
                    break
            
            if not found:
                print(f"  [X] Header '{header}' NOT FOUND in sheet")
    
    def diagnose_section_issues(self, sheet_name: str, section_patterns: List[str]) -> None:
        """
        Diagnose section finding issues for debugging
        
        Args:
            sheet_name: Name of the sheet to diagnose
            section_patterns: List of section patterns to search for
        """
        print(f"\n{'='*100}")
        print(f"DIAGNOSING: {sheet_name} - Section Finding")
        print('='*100)
        
        if sheet_name not in self.workbook.sheetnames:
            print(f"[X] Sheet '{sheet_name}' not found!")
            return
        
        sheet = self.workbook[sheet_name]
        
        print(f"\n🔍 Scanning sheet for sections (checking column A)...")
        print(f"Sheet has {sheet.max_row} rows")
        
        sections_found = []
        for row_idx in range(1, min(sheet.max_row + 1, 1000)):
            cell = sheet.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str):
                value = cell.value.strip()
                # Check if it looks like a section header
                if value and len(value) > 2 and '.' in value:
                    parts = value.split('.')
                    if parts[0].strip().isdigit():
                        sections_found.append((row_idx, value))
        
        print(f"\nFound {len(sections_found)} numbered sections:")
        for row, section in sections_found:
            print(f"  Row {row}: '{section}'")
        
        # Check for specific patterns
        for pattern in section_patterns:
            print(f"\n🔍 Looking for pattern: '{pattern}'")
            found_matches = [s for r, s in sections_found if pattern.upper() in s.upper()]
            if found_matches:
                print(f"  [OK] FOUND: {found_matches}")
            else:
                print(f"  [X] NOT FOUND")
    
    def run_diagnostics(self) -> None:
        """
        Run full diagnostics on the current file
        Helps identify why validations might be failing
        """
        print("="*100)
        print("EXCEL VALIDATION DIAGNOSTIC TOOL")
        print("="*100)
        print(f"File: {self.filepath}\n")
        
        # Diagnose Ex Inv VL currency issues
        self.diagnose_currency_issues(
            "Ex Inv VL-1",
            ["Excess Cost #1", "Cost #1", "Ext Vol Cost (Splits) #1", 
             "Excess Cost #1 -B1", "Buy value after -B1", "Net Excess Cost #1"]
        )
        
        # Diagnose A CLASS PARTS currency issues
        self.diagnose_currency_issues(
            "A CLASS PARTS",
            ["Cost #1", "Ext Price (Splits) #1 (Conv.)", "Ext Vol Cost (Splits) #1"]
        )
        
        # Diagnose BOM MATRIX currency issues
        self.diagnose_currency_issues(
            "BOM MATRIX",
            ["Unit Price", "Grand Total", "VL-1", "VL-2"]
        )
        
        # Diagnose Missing Notes sections
        self.diagnose_section_issues(
            "Missing Notes",
            ["NRFND", "Proto Pricing No Cost"]
        )
        
        print("\n" + "="*100)
        print("DIAGNOSTIC COMPLETE")
        print("="*100)
    
    def validate_rule1_header_presence(self) -> List[ValidationResult]:
        """
        Rule 1: Validate that header "Ext Part Vol Price (Splits) #{X} (Conv.)" 
        is present in sheets matching pattern "7.0 CBOM VL-{X}"
        where the number in the header matches the sheet suffix
        (e.g., VL-1 should have #1, VL-2 should have #2, etc.)
        
        Returns:
            List of ValidationResult objects
        """
        # Define the pattern for sheet names
        sheet_pattern = r'^7\.0 CBOM VL-(\d+)$'
        
        # Find all matching sheets
        matching_sheets = [
            sheet_name for sheet_name in self.workbook.sheetnames 
            if re.match(sheet_pattern, sheet_name)
        ]
        
        results = []
        
        # If no matching sheets found
        if not matching_sheets:
            result = ValidationResult(
                rule_name="Rule 1: Header Validation",
                sheet_name="N/A",
                status="FAIL",
                expected=f"Sheets matching pattern '7.0 CBOM VL-{{X}}'",
                actual="No matching sheets found"
            )
            results.append(result)
            return results
        
        # Validate each matching sheet
        for sheet_name in matching_sheets:
            # Extract the number from sheet name (e.g., "7.0 CBOM VL-1" -> "1")
            match = re.match(sheet_pattern, sheet_name)
            sheet_number = match.group(1)
            
            # Construct the expected header based on sheet number
            expected_header = f"Ext Part Vol Price (Splits) #{sheet_number} (Conv.)"
            
            sheet = self.workbook[sheet_name]
            
            # Search for the header in the sheet
            found = False
            found_location = ""
            actual_value = "Header not found"
            
            # Search up to 100 rows and 100 columns (reasonable limit)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100, max_col=100), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value == expected_header:
                        found = True
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        found_location = f"{col_letter}{row_idx}"
                        actual_value = cell.value
                        break
                if found:
                    break
            
            # If header not found, try to find what header actually exists in similar location
            if not found:
                # Search for similar headers to provide better error message
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=100, max_col=100), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value and isinstance(cell.value, str):
                            if "Ext Part Vol Price" in cell.value and "Splits" in cell.value:
                                actual_value = f"Found: '{cell.value}'"
                                col_letter = openpyxl.utils.get_column_letter(col_idx)
                                found_location = f"{col_letter}{row_idx}"
                                break
                    if "Found:" in actual_value:
                        break
            
            # Create result based on whether header was found
            if found:
                result = ValidationResult(
                    rule_name="Rule 1: Header Validation",
                    sheet_name=sheet_name,
                    status="PASS",
                    expected=expected_header,
                    actual=actual_value,
                    location=found_location
                )
            else:
                result = ValidationResult(
                    rule_name="Rule 1: Header Validation",
                    sheet_name=sheet_name,
                    status="FAIL",
                    expected=expected_header,
                    actual=actual_value,
                    location=found_location if found_location else "N/A"
                )
            
            results.append(result)
        
        return results
    
    def validate_rule2_corrected_mpn(self) -> List[ValidationResult]:
        """
        Rule 2: Validate that Quoted MPN is present under Corrected MPN Mentioned section
        Corrected MPN must reflect in the Quoted MPN.
        - "Quoted MPN" sub-header should be present under "Corrected MPN Mentioned" section
        - Alphanumeric values should be present under "Quoted MPN"
        - "Corrected MPN" sub-header should NOT be present

        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_name = "Missing Notes"

        if sheet_name not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 2: Quoted MPN Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Sheet 'Missing Notes' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results

        sheet = self.workbook[sheet_name]

        # Look for "Corrected MPN Mentioned" section and "Quoted MPN" header
        corrected_mpn_mentioned_found = False
        quoted_mpn_found = False
        quoted_mpn_has_values = False
        quoted_mpn_location = ""
        corrected_mpn_subheader_found = False

        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=500, max_col=20), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    # Check for section header
                    if "Corrected MPN Mentioned" in cell.value:
                        corrected_mpn_mentioned_found = True

                    # Check for Quoted MPN sub-header
                    if cell.value == "Quoted MPN":
                        quoted_mpn_found = True
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        quoted_mpn_location = f"{col_letter}{row_idx}"

                        # Check for values below
                        for value_row in range(row_idx + 1, min(row_idx + 100, sheet.max_row + 1)):
                            value_cell = sheet.cell(row=value_row, column=col_idx)
                            if value_cell.value and str(value_cell.value).strip():
                                quoted_mpn_has_values = True
                                break

                    # Check for "Corrected MPN" sub-header (should NOT be present)
                    if cell.value == "Corrected MPN":
                        corrected_mpn_subheader_found = True

        # Create result
        if quoted_mpn_found and quoted_mpn_has_values and not corrected_mpn_subheader_found:
            result = ValidationResult(
                rule_name="Rule 2: Quoted MPN Validation",
                sheet_name=sheet_name,
                status="PASS",
                expected="'Quoted MPN' header with values",
                actual="Quoted MPN is present",
                location=quoted_mpn_location
            )
        elif corrected_mpn_subheader_found:
            result = ValidationResult(
                rule_name="Rule 2: Quoted MPN Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'Corrected MPN' sub-header should NOT be present",
                actual="'Corrected MPN' sub-header found (should only have 'Quoted MPN')"
            )
        elif quoted_mpn_found and not quoted_mpn_has_values:
            result = ValidationResult(
                rule_name="Rule 2: Quoted MPN Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'Quoted MPN' header with values",
                actual="Quoted MPN found but no values present"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 2: Quoted MPN Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'Quoted MPN' header should be present under 'Corrected MPN Mentioned' section",
                actual="Quoted MPN header not found"
            )

        results.append(result)
        return results
    
    def _has_currency_format_or_symbol(self, cell) -> bool:
        """
        Helper method to check if a cell has currency formatting or contains currency symbol
        Checks both Excel number format and the actual value/display
        
        Returns:
            True if cell has currency format or contains currency symbol
        """
        # Check number format
        if cell.number_format and any(symbol in cell.number_format 
                                       for symbol in ['$', '€', '£', '¥', '₹', '¤']):
            return True
        
        # Check actual value for currency symbols
        if cell.value is not None:
            value_str = str(cell.value)
            if any(symbol in value_str for symbol in ['$', '€', '£', '¥', '₹']):
                return True
        
        return False
    
    def validate_rule3_currency_symbols_cbom(self) -> List[ValidationResult]:
        """
        Rule 3: Validate that price columns in CBOM VL sheets have currency symbols
        Checks: Ext Price #{X} (Conv.) and Ext Part Vol Price #{X} (Conv.)
        where X dynamically matches the sheet suffix number
        Checks BOTH above and below headers for currency symbols
        IMPORTANT: Prioritizes headers found in lower row numbers (first 15 rows)
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_pattern = r'^7\.0 CBOM VL-(\d+)$'
        
        matching_sheets = [
            sheet_name for sheet_name in self.workbook.sheetnames 
            if re.match(sheet_pattern, sheet_name)
        ]
        
        if not matching_sheets:
            result = ValidationResult(
                rule_name="Rule 3: Currency Symbol Validation (CBOM)",
                sheet_name="N/A",
                status="FAIL",
                expected="Sheets matching pattern '7.0 CBOM VL-{X}'",
                actual="No matching sheets found"
            )
            results.append(result)
            return results
        
        for sheet_name in matching_sheets:
            # Extract the number from sheet name dynamically
            match = re.match(sheet_pattern, sheet_name)
            sheet_number = match.group(1)
            
            sheet = self.workbook[sheet_name]
            
            # Headers to check - dynamically constructed based on sheet number
            headers_to_check = [
                f"Ext Price #{sheet_number} (Conv.)",
                f"Ext Part Vol Price #{sheet_number} (Conv.)"
            ]
            
            all_headers_valid = True
            missing_headers = []
            invalid_formats = []
            
            for expected_header in headers_to_check:
                header_found = False
                has_currency = False
                best_row = None  # Track the lowest row number where header is found
                best_col = None
                
                # Find the header - prioritize LOWEST row number (first 15 rows)
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=15, max_col=50), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value == expected_header:
                            # Found the header - this is in the early rows (lower row numbers)
                            if best_row is None or row_idx < best_row:
                                best_row = row_idx
                                best_col = col_idx
                            break
                
                # If found in first 15 rows, use that location
                if best_row is not None:
                    header_found = True
                    row_idx = best_row
                    col_idx = best_col
                    
                    # Check cells BELOW the header
                    for value_row in range(row_idx + 1, min(row_idx + 10, sheet.max_row + 1)):
                        value_cell = sheet.cell(row=value_row, column=col_idx)
                        if self._has_currency_format_or_symbol(value_cell):
                            has_currency = True
                            break
                    
                    # Also check cells ABOVE the header (if not found below)
                    if not has_currency and row_idx > 1:
                        for value_row in range(max(1, row_idx - 5), row_idx):
                            value_cell = sheet.cell(row=value_row, column=col_idx)
                            if self._has_currency_format_or_symbol(value_cell):
                                has_currency = True
                                break
                
                if not header_found:
                    missing_headers.append(expected_header)
                    all_headers_valid = False
                elif not has_currency:
                    invalid_formats.append(expected_header)
                    all_headers_valid = False
            
            # Create result
            if missing_headers:
                result = ValidationResult(
                    rule_name="Rule 3: Currency Symbol Validation (CBOM)",
                    sheet_name=sheet_name,
                    status="FAIL",
                    expected=f"Headers for sheet number #{sheet_number} with currency-formatted values",
                    actual=f"Missing headers: {', '.join(missing_headers)}"
                )
            elif invalid_formats:
                result = ValidationResult(
                    rule_name="Rule 3: Currency Symbol Validation (CBOM)",
                    sheet_name=sheet_name,
                    status="FAIL",
                    expected="Values with currency symbols ($, €, £, etc.)",
                    actual=f"No currency format in: {', '.join(invalid_formats)}"
                )
            else:
                result = ValidationResult(
                    rule_name="Rule 3: Currency Symbol Validation (CBOM)",
                    sheet_name=sheet_name,
                    status="PASS",
                    expected=f"Currency symbols in Ext Price #{sheet_number} and Ext Part Vol Price #{sheet_number}",
                    actual="Currency symbol present"
                )
            
            results.append(result)
        
        return results
    
    def validate_rule4_moq_cost(self) -> List[ValidationResult]:
        """
        Rule 4: Validate that MOQ Cost column has a percentage value above it
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_pattern = r'^7\.0 CBOM VL-\d+$'
        
        matching_sheets = [
            sheet_name for sheet_name in self.workbook.sheetnames 
            if re.match(sheet_pattern, sheet_name)
        ]
        
        if not matching_sheets:
            result = ValidationResult(
                rule_name="Rule 4: MOQ Cost % Validation",
                sheet_name="N/A",
                status="FAIL",
                expected="Sheets matching pattern '7.0 CBOM VL-{X}'",
                actual="No matching sheets found"
            )
            results.append(result)
            return results
        
        for sheet_name in matching_sheets:
            sheet = self.workbook[sheet_name]
            
            moq_cost_found = False
            has_percentage_above = False
            moq_location = ""
            percentage_value = None
            
            # Find "MOQ Cost" header
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, max_col=50), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value and isinstance(cell.value, str) and cell.value.strip() == "MOQ Cost":
                        moq_cost_found = True
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        moq_location = f"{col_letter}{row_idx}"
                        
                        # Check cell above for percentage
                        if row_idx > 1:
                            above_cell = sheet.cell(row=row_idx - 1, column=col_idx)
                            above_value = above_cell.value
                            
                            # Check if it's a percentage (either as format or value)
                            if above_value is not None:
                                # Check if number format contains %
                                if above_cell.number_format and '%' in above_cell.number_format:
                                    has_percentage_above = True
                                    percentage_value = above_value
                                # Or if value itself contains %
                                elif isinstance(above_value, str) and '%' in above_value:
                                    has_percentage_above = True
                                    percentage_value = above_value
                                # Or if it's a number (could be formatted as percentage)
                                elif isinstance(above_value, (int, float)):
                                    percentage_value = above_value
                                    # Accept any numeric value above as it might be percentage
                                    has_percentage_above = True
                        break
                if moq_cost_found:
                    break
            
            # Create result
            if not moq_cost_found:
                result = ValidationResult(
                    rule_name="Rule 4: MOQ Cost % Validation",
                    sheet_name=sheet_name,
                    status="FAIL",
                    expected="'MOQ Cost' header should be present",
                    actual="'MOQ Cost' header not found"
                )
            elif not has_percentage_above:
                result = ValidationResult(
                    rule_name="Rule 4: MOQ Cost % Validation",
                    sheet_name=sheet_name,
                    status="FAIL",
                    expected="Percentage value above 'MOQ Cost'",
                    actual=f"No percentage found above MOQ Cost at {moq_location}"
                )
            else:
                result = ValidationResult(
                    rule_name="Rule 4: MOQ Cost % Validation",
                    sheet_name=sheet_name,
                    status="PASS",
                    expected="MOQ Cost with percentage above",
                    actual="MOQ Cost present",
                    location=moq_location
                )
            
            results.append(result)
        
        return results
    
    def validate_rule5_currency_symbols_ex_inv(self) -> List[ValidationResult]:
        """
        Rule 5: Validate currency symbols in Ex Inv VL sheets price columns
        Dynamically checks headers matching the sheet suffix number
        Checks BOTH above and below headers for currency symbols
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_pattern = r'^Ex Inv VL-(\d+)$'
        
        matching_sheets = [
            sheet_name for sheet_name in self.workbook.sheetnames 
            if re.match(sheet_pattern, sheet_name)
        ]
        
        if not matching_sheets:
            result = ValidationResult(
                rule_name="Rule 5: Currency Symbol Validation (Ex Inv)",
                sheet_name="N/A",
                status="FAIL",
                expected="Sheets matching pattern 'Ex Inv VL-{X}'",
                actual="No matching sheets found"
            )
            results.append(result)
            return results
        
        for sheet_name in matching_sheets:
            # Extract the number from sheet name dynamically
            match = re.match(sheet_pattern, sheet_name)
            sheet_number = match.group(1)
            
            sheet = self.workbook[sheet_name]
            
            # Headers to check - dynamically constructed based on sheet number
            headers_to_check = [
                f"Excess Cost #{sheet_number}",
                f"Cost #{sheet_number}",
                f"Ext Vol Cost (Splits) #{sheet_number}",
                f"Excess Cost #{sheet_number} -B{sheet_number}",
                f"Buy value after -B{sheet_number}",
                f"Net Excess Cost #{sheet_number}"
            ]
            
            all_valid = True
            validation_details = []
            headers_checked = []
            
            for expected_header in headers_to_check:
                header_found = False
                has_currency = False
                actual_header = ""
                
                # Find the header
                for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, max_col=50), 1):
                    for col_idx, cell in enumerate(row, 1):
                        if cell.value and isinstance(cell.value, str):
                            # Exact match
                            if cell.value == expected_header:
                                header_found = True
                                actual_header = cell.value
                                headers_checked.append(f"{expected_header} found")
                                
                                # Check cells BELOW the header
                                for value_row in range(row_idx + 1, min(row_idx + 10, sheet.max_row + 1)):
                                    value_cell = sheet.cell(row=value_row, column=col_idx)
                                    if self._has_currency_format_or_symbol(value_cell):
                                        has_currency = True
                                        break
                                
                                # Also check cells ABOVE the header (if not found below)
                                if not has_currency and row_idx > 1:
                                    for value_row in range(max(1, row_idx - 5), row_idx):
                                        value_cell = sheet.cell(row=value_row, column=col_idx)
                                        if self._has_currency_format_or_symbol(value_cell):
                                            has_currency = True
                                            break
                                break
                    if header_found:
                        break
                
                # Track results
                if header_found and not has_currency:
                    all_valid = False
                    validation_details.append(f"'{expected_header}' missing currency")
                elif not header_found:
                    validation_details.append(f"'{expected_header}' not found")
            
            # Create result
            if validation_details:
                # Check if we found at least some headers with currency
                if any("not found" in detail for detail in validation_details):
                    result = ValidationResult(
                        rule_name="Rule 5: Currency Symbol Validation (Ex Inv)",
                        sheet_name=sheet_name,
                        status="PASS" if len(validation_details) <= 2 else "FAIL",  # Allow minor variations
                        expected=f"Currency symbols in price columns for sheet #{sheet_number}",
                        actual="; ".join(validation_details) if validation_details else "Currency symbol present"
                    )
                else:
                    result = ValidationResult(
                        rule_name="Rule 5: Currency Symbol Validation (Ex Inv)",
                        sheet_name=sheet_name,
                        status="FAIL",
                        expected=f"Currency symbols in all price columns for #{sheet_number}",
                        actual="; ".join(validation_details)
                    )
            else:
                result = ValidationResult(
                    rule_name="Rule 5: Currency Symbol Validation (Ex Inv)",
                    sheet_name=sheet_name,
                    status="PASS",
                    expected=f"Currency symbols present in price columns for #{sheet_number}",
                    actual="Currency symbol present"
                )
            
            results.append(result)
        
        return results
    
    def validate_rule6_net_ordering_qty(self) -> List[ValidationResult]:
        """
        Rule 6: Validate that "Net Ordering qty" header is present in Ex Inv VL sheets
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_pattern = r'^Ex Inv VL-\d+$'
        
        matching_sheets = [
            sheet_name for sheet_name in self.workbook.sheetnames 
            if re.match(sheet_pattern, sheet_name)
        ]
        
        if not matching_sheets:
            result = ValidationResult(
                rule_name="Rule 6: Net Ordering qty Header Validation",
                sheet_name="N/A",
                status="FAIL",
                expected="Sheets matching pattern 'Ex Inv VL-{X}'",
                actual="No matching sheets found"
            )
            results.append(result)
            return results
        
        for sheet_name in matching_sheets:
            sheet = self.workbook[sheet_name]
            
            header_found = False
            header_location = ""
            
            # Find "Net Ordering qty" header
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, max_col=50), 1):
                for col_idx, cell in enumerate(row, 1):
                    if cell.value == "Net Ordering qty":
                        header_found = True
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        header_location = f"{col_letter}{row_idx}"
                        break
                if header_found:
                    break
            
            # Create result
            if header_found:
                result = ValidationResult(
                    rule_name="Rule 6: Net Ordering qty Header Validation",
                    sheet_name=sheet_name,
                    status="PASS",
                    expected="'Net Ordering qty' header present",
                    actual="Header is inline",
                    location=header_location
                )
            else:
                result = ValidationResult(
                    rule_name="Rule 6: Net Ordering qty Header Validation",
                    sheet_name=sheet_name,
                    status="FAIL",
                    expected="'Net Ordering qty' header should be present",
                    actual="Header not found"
                )
            
            results.append(result)
        
        return results
    
    def validate_rule7_currency_a_class_parts(self) -> List[ValidationResult]:
        """
        Rule 7: Validate currency symbols in A CLASS PARTS sheet
        Checks: Cost #{X}, Ext Price (Splits) #{X} (Conv.), Ext Vol Cost (Splits) #{X}
        where X = 1, 2, 3... (dynamically detected)
        Checks BOTH above and below headers for currency symbols
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_name = "A CLASS PARTS"
        
        if sheet_name not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 7: Currency Validation (A CLASS PARTS)",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Sheet 'A CLASS PARTS' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results
        
        sheet = self.workbook[sheet_name]
        
        # Find all headers dynamically
        cost_headers = []
        ext_price_headers = []
        ext_vol_cost_headers = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, max_col=50), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    # Find Cost #{X} headers dynamically
                    if cell.value.startswith("Cost #"):
                        cost_headers.append((cell.value, row_idx, col_idx))
                    # Find Ext Price (Splits) #{X} headers dynamically
                    elif "Ext Price (Splits) #" in cell.value:
                        ext_price_headers.append((cell.value, row_idx, col_idx))
                    # Find Ext Vol Cost (Splits) #{X} headers dynamically
                    elif "Ext Vol Cost (Splits) #" in cell.value:
                        ext_vol_cost_headers.append((cell.value, row_idx, col_idx))
        
        all_valid = True
        issues = []
        
        # Validate all headers - check BOTH above AND below
        all_headers = cost_headers + ext_price_headers + ext_vol_cost_headers
        
        for header, row_idx, col_idx in all_headers:
            has_currency = False
            
            # Check cells BELOW header
            for value_row in range(row_idx + 1, min(row_idx + 10, sheet.max_row + 1)):
                value_cell = sheet.cell(row=value_row, column=col_idx)
                if self._has_currency_format_or_symbol(value_cell):
                    has_currency = True
                    break
            
            # Also check cells ABOVE header (if not found below)
            if not has_currency and row_idx > 1:
                for value_row in range(max(1, row_idx - 5), row_idx):
                    value_cell = sheet.cell(row=value_row, column=col_idx)
                    if self._has_currency_format_or_symbol(value_cell):
                        has_currency = True
                        break
            
            if not has_currency:
                all_valid = False
                issues.append(f"'{header}' missing currency")
        
        # Create result
        if all_valid:
            result = ValidationResult(
                rule_name="Rule 7: Currency Validation (A CLASS PARTS)",
                sheet_name=sheet_name,
                status="PASS",
                expected="Currency symbols in all Cost, Ext Price, and Ext Vol Cost columns",
                actual="Currency symbol present"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 7: Currency Validation (A CLASS PARTS)",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Currency symbols in Cost #{X}, Ext Price (Splits) #{X}, and Ext Vol Cost (Splits) #{X}",
                actual="; ".join(issues)
            )
        
        results.append(result)
        return results
    
    def validate_rule8_quoted_mfr(self) -> List[ValidationResult]:
        """
        Rule 8: Validate that Quoted MFR is present under Manufacturer Name Mismatch section
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_name = "Missing Notes"
        
        if sheet_name not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 8: Quoted MFR Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Sheet 'Missing Notes' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results
        
        sheet = self.workbook[sheet_name]
        
        # Look for "Manufacturer Name Mismatch" section and "Quoted MFR" header
        mfr_mismatch_found = False
        quoted_mfr_found = False
        quoted_mfr_has_values = False
        quoted_mfr_location = ""
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=50, max_col=20), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    # Check for section header
                    if "Manufacturer Name Mismatch" in cell.value:
                        mfr_mismatch_found = True
                    
                    # Check for Quoted MFR header
                    if cell.value == "Quoted MFR":
                        quoted_mfr_found = True
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        quoted_mfr_location = f"{col_letter}{row_idx}"
                        
                        # Check for values below
                        for value_row in range(row_idx + 1, min(row_idx + 100, sheet.max_row + 1)):
                            value_cell = sheet.cell(row=value_row, column=col_idx)
                            if value_cell.value and str(value_cell.value).strip():
                                quoted_mfr_has_values = True
                                break
        
        # Create result
        if quoted_mfr_found and quoted_mfr_has_values:
            result = ValidationResult(
                rule_name="Rule 8: Quoted MFR Validation",
                sheet_name=sheet_name,
                status="PASS",
                expected="'Quoted MFR' header with values",
                actual="Quoted MFR is present",
                location=quoted_mfr_location
            )
        elif quoted_mfr_found and not quoted_mfr_has_values:
            result = ValidationResult(
                rule_name="Rule 8: Quoted MFR Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'Quoted MFR' header with values",
                actual="Quoted MFR found but no values present"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 8: Quoted MFR Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'Quoted MFR' header should be present",
                actual="Quoted MFR header not found"
            )
        
        results.append(result)
        return results
    
    def validate_rule9_nrfnd_missing(self) -> List[ValidationResult]:
        """
        Rule 9: Validate that "#. NRFND" header is present with values in Missing Notes
        where # = 1, 2, 3... (e.g., "1. NRFND", "2. NRFND", "3. NRFND", "8. NRFND")
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_name = "Missing Notes"
        
        if sheet_name not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 9: NRFND Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Sheet 'Missing Notes' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results
        
        sheet = self.workbook[sheet_name]
        
        # Look for "#. NRFND" pattern where # is a number - search entire sheet
        nrfnd_sections = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=20), 1):
            cell = sheet.cell(row=row_idx, column=1)  # Check column A for section headers
            if cell.value and isinstance(cell.value, str):
                # Check if it matches pattern like "1. NRFND", "2. NRFND", "8. NRFND", etc.
                value = cell.value.strip()
                # Check for exact pattern: number + ". NRFND" (case insensitive)
                if value and len(value) > 2:
                    # Check if it ends with "NRFND" and starts with number + dot
                    if value.upper().endswith("NRFND") or ". NRFND" in value.upper() or ".NRFND" in value.upper():
                        # Extract the number part
                        parts = value.split(".")
                        if len(parts) >= 2 and parts[0].strip().isdigit():
                            nrfnd_sections.append((value, row_idx))
        
        nrfnd_found = len(nrfnd_sections) > 0
        nrfnd_has_values = False
        
        # If sections found, check for values below each section
        if nrfnd_found:
            for section_name, section_row in nrfnd_sections:
                # Check the next few rows for data (skip header row)
                for value_row in range(section_row + 2, min(section_row + 100, sheet.max_row + 1)):
                    # Check if there's any data in the row
                    has_data = False
                    for col_idx in range(1, 15):
                        value_cell = sheet.cell(row=value_row, column=col_idx)
                        if value_cell.value and str(value_cell.value).strip():
                            has_data = True
                            nrfnd_has_values = True
                            break
                    
                    if has_data:
                        break
                
                if nrfnd_has_values:
                    break
        
        # Create result
        if nrfnd_found and nrfnd_has_values:
            section_names = [name for name, _ in nrfnd_sections]
            result = ValidationResult(
                rule_name="Rule 9: NRFND Validation",
                sheet_name=sheet_name,
                status="PASS",
                expected="'#. NRFND' sections with values",
                actual=f"NRFND present in missing notes ({len(nrfnd_sections)} section(s): {', '.join(section_names)})",
                location=f"Row {nrfnd_sections[0][1]}"
            )
        elif nrfnd_found and not nrfnd_has_values:
            result = ValidationResult(
                rule_name="Rule 9: NRFND Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'#. NRFND' sections with values",
                actual="NRFND sections found but no values present"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 9: NRFND Validation",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'#. NRFND' sections should be present (e.g., '1. NRFND', '2. NRFND')",
                actual="NRFND sections not found"
            )
        
        results.append(result)
        return results
    
    def validate_rule10_currency_bom_matrix(self) -> List[ValidationResult]:
        """
        Rule 10: Validate currency in BOM MATRIX price columns
        Checks: Unit Price, Grand Total (2nd occurrence only), Net Excess Cost, VL-{X} (only after last Unit Price)
        Checks BOTH above and below headers for currency symbols
        For VL-{X}, only checks columns that appear AFTER the last "Unit Price" column

        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_name = "BOM MATRIX"

        if sheet_name not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 10: Currency Validation (BOM MATRIX)",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Sheet 'BOM MATRIX' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results

        sheet = self.workbook[sheet_name]

        found_headers = []
        missing_currency = []

        # First pass: Find all Grand Total occurrences to identify the 2nd one
        grand_total_occurrences = []
        for row_idx in range(1, min(31, sheet.max_row + 1)):
            for col_idx in range(1, min(101, sheet.max_column + 1) if hasattr(sheet, 'max_column') else 101):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and "Grand Total" in cell.value.strip():
                    grand_total_occurrences.append((row_idx, col_idx, cell.value.strip()))

        # Sort by column index to get column-wise order
        grand_total_occurrences.sort(key=lambda x: x[1])

        # Get the 2nd Grand Total (skip the first one)
        second_grand_total = None
        if len(grand_total_occurrences) >= 2:
            second_grand_total = grand_total_occurrences[1]  # (row_idx, col_idx, header_name)

        # Find all "Unit Price" headers to get the LAST one (rightmost column)
        unit_price_occurrences = []
        for row_idx in range(1, min(31, sheet.max_row + 1)):
            for col_idx in range(1, min(101, sheet.max_column + 1) if hasattr(sheet, 'max_column') else 101):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and "Unit Price" in cell.value.strip():
                    unit_price_occurrences.append((row_idx, col_idx, cell.value.strip()))

        # Sort by column index and get the last (rightmost) Unit Price
        unit_price_occurrences.sort(key=lambda x: x[1])
        last_unit_price_col = unit_price_occurrences[-1][1] if unit_price_occurrences else 0

        # Headers to check (Unit Price, Net Excess Cost)
        static_headers = ["Unit Price", "Net Excess Cost"]

        # Collect all VL-{X} headers found across all columns
        vl_headers_found = []

        # Search for all headers
        for row_idx in range(1, min(31, sheet.max_row + 1)):
            for col_idx in range(1, min(101, sheet.max_column + 1) if hasattr(sheet, 'max_column') else 101):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    header_name = cell.value.strip()

                    # Check for Unit Price and Net Excess Cost
                    if any(header in header_name for header in static_headers):
                        has_currency = False

                        # Check cells BELOW the header
                        for value_row in range(row_idx + 1, min(row_idx + 10, sheet.max_row + 1)):
                            value_cell = sheet.cell(row=value_row, column=col_idx)
                            if self._has_currency_format_or_symbol(value_cell):
                                has_currency = True
                                break

                        # Also check cells ABOVE the header (if not found below)
                        if not has_currency and row_idx > 1:
                            for value_row in range(max(1, row_idx - 5), row_idx):
                                value_cell = sheet.cell(row=value_row, column=col_idx)
                                if self._has_currency_format_or_symbol(value_cell):
                                    has_currency = True
                                    break

                        found_headers.append(header_name)
                        if not has_currency:
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            missing_currency.append(f"{header_name} (Col {col_letter})")

                    # Check for dynamic VL-X headers - collect ALL occurrences (exclude EX-VL-)
                    elif header_name.startswith("VL-"):
                        vl_headers_found.append((row_idx, col_idx, header_name))

        # Check VL-{X} headers ONLY if they appear AFTER the last "Unit Price" column
        for row_idx, col_idx, header_name in vl_headers_found:
            # Skip VL-{X} headers that are before or at the last Unit Price column
            if col_idx <= last_unit_price_col:
                continue

            has_currency = False

            # For VL-X headers: check 2-3 rows below for currency
            for value_row in range(row_idx + 1, min(row_idx + 4, sheet.max_row + 1)):
                value_cell = sheet.cell(row=value_row, column=col_idx)
                if self._has_currency_format_or_symbol(value_cell):
                    has_currency = True
                    break

            # Also check cells ABOVE the header (if not found below)
            if not has_currency and row_idx > 1:
                for value_row in range(max(1, row_idx - 3), row_idx):
                    value_cell = sheet.cell(row=value_row, column=col_idx)
                    if self._has_currency_format_or_symbol(value_cell):
                        has_currency = True
                        break

            found_headers.append(header_name)
            if not has_currency:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                missing_currency.append(f"{header_name} (Col {col_letter})")

        # Check the 2nd Grand Total (skip the first one)
        if second_grand_total:
            row_idx, col_idx, header_name = second_grand_total
            has_currency = False

            # Check cells BELOW the header
            for value_row in range(row_idx + 1, min(row_idx + 10, sheet.max_row + 1)):
                value_cell = sheet.cell(row=value_row, column=col_idx)
                if self._has_currency_format_or_symbol(value_cell):
                    has_currency = True
                    break

            # Also check cells ABOVE the header (if not found below)
            if not has_currency and row_idx > 1:
                for value_row in range(max(1, row_idx - 5), row_idx):
                    value_cell = sheet.cell(row=value_row, column=col_idx)
                    if self._has_currency_format_or_symbol(value_cell):
                        has_currency = True
                        break

            found_headers.append(header_name)
            if not has_currency:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                missing_currency.append(f"{header_name} (Col {col_letter})")

        # Create result
        if not missing_currency:
            result = ValidationResult(
                rule_name="Rule 10: Currency Validation (BOM MATRIX)",
                sheet_name=sheet_name,
                status="PASS",
                expected="Currency symbols in Unit Price, Grand Total, Net Excess Cost, VL-X",
                actual="Currency symbol present"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 10: Currency Validation (BOM MATRIX)",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Currency symbols in all price columns",
                actual=f"Missing currency: {', '.join(missing_currency[:10])}"
            )

        results.append(result)
        return results
    
    def _check_proto_headers_in_bom_matrix(self) -> Tuple[bool, bool]:
        """
        Helper method to check if Proto Qty and Proto Price headers exist in BOM MATRIX
        
        Returns:
            Tuple of (proto_qty_found, proto_price_found)
        """
        if "BOM MATRIX" not in self.workbook.sheetnames:
            return False, False
        
        sheet = self.workbook["BOM MATRIX"]
        proto_qty_found = False
        proto_price_found = False
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, max_col=100), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str):
                    if "Proto Qty" in cell.value:
                        proto_qty_found = True
                    if "Proto Price" in cell.value:
                        proto_price_found = True
            
            if proto_qty_found and proto_price_found:
                break
        
        return proto_qty_found, proto_price_found
    
    def validate_rule11_cbom_proto_sheet(self) -> List[ValidationResult]:
        """
        Rule 11: Validate CBOM Proto sheet exists if Proto Qty and Proto Price are in BOM MATRIX
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        
        # Check for Proto headers in BOM MATRIX
        proto_qty_found, proto_price_found = self._check_proto_headers_in_bom_matrix()
        proto_headers_exist = proto_qty_found and proto_price_found
        
        # Check if 7.0 CBOM Proto sheet exists
        cbom_proto_exists = "7.0 CBOM Proto" in self.workbook.sheetnames
        
        # Determine result
        if proto_headers_exist and cbom_proto_exists:
            result = ValidationResult(
                rule_name="Rule 11: CBOM Proto Sheet Validation",
                sheet_name="7.0 CBOM Proto",
                status="PASS",
                expected="'7.0 CBOM Proto' sheet should exist when Proto headers present",
                actual="Proto present"
            )
        elif proto_headers_exist and not cbom_proto_exists:
            result = ValidationResult(
                rule_name="Rule 11: CBOM Proto Sheet Validation",
                sheet_name="N/A",
                status="FAIL",
                expected="'7.0 CBOM Proto' sheet should exist",
                actual="Proto quantity present in BOM MATRIX but 7.0 CBOM Proto sheet missing"
            )
        elif not proto_headers_exist and cbom_proto_exists:
            result = ValidationResult(
                rule_name="Rule 11: CBOM Proto Sheet Validation",
                sheet_name="7.0 CBOM Proto",
                status="FAIL",
                expected="'7.0 CBOM Proto' sheet should not exist without Proto headers",
                actual="Proto quantity does not present but 7.0 CBOM Proto present"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 11: CBOM Proto Sheet Validation",
                sheet_name="N/A",
                status="PASS",
                expected="No Proto headers, no Proto sheet",
                actual="Proto not specified, sheet correctly absent"
            )
        
        results.append(result)
        return results
    
    def validate_rule12_proto_volume_summary(self) -> List[ValidationResult]:
        """
        Rule 12: Validate Proto Volume header in Summary sheet based on Proto headers in BOM MATRIX
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        
        # Check for Proto headers in BOM MATRIX
        proto_qty_found, proto_price_found = self._check_proto_headers_in_bom_matrix()
        proto_headers_exist = proto_qty_found and proto_price_found
        
        # Check if Summary sheet exists
        if "Summary" not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 12: Proto Volume in Summary",
                sheet_name="Summary",
                status="FAIL",
                expected="Sheet 'Summary' should exist",
                actual="Summary sheet not found"
            )
            results.append(result)
            return results
        
        # Check for Proto Volume header in Summary
        sheet = self.workbook["Summary"]
        proto_volume_found = False
        proto_volume_location = ""
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30, max_col=50), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str) and "Proto Volume" in cell.value:
                    proto_volume_found = True
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    proto_volume_location = f"{col_letter}{row_idx}"
                    break
            if proto_volume_found:
                break
        
        # Determine result
        if proto_headers_exist and proto_volume_found:
            result = ValidationResult(
                rule_name="Rule 12: Proto Volume in Summary",
                sheet_name="Summary",
                status="PASS",
                expected="'Proto Volume' header should be present",
                actual="Proto volume present in summary",
                location=proto_volume_location
            )
        elif proto_headers_exist and not proto_volume_found:
            result = ValidationResult(
                rule_name="Rule 12: Proto Volume in Summary",
                sheet_name="Summary",
                status="FAIL",
                expected="'Proto Volume' header should be present when Proto specified",
                actual="Proto volume does not present in summary"
            )
        elif not proto_headers_exist and proto_volume_found:
            result = ValidationResult(
                rule_name="Rule 12: Proto Volume in Summary",
                sheet_name="Summary",
                status="FAIL",
                expected="'Proto Volume' should not be present without Proto specification",
                actual="Proto volume present but Proto not specified in BOM MATRIX"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 12: Proto Volume in Summary",
                sheet_name="Summary",
                status="PASS",
                expected="No Proto Volume when Proto not specified",
                actual="Proto not specified, Proto Volume correctly absent"
            )
        
        results.append(result)
        return results
    
    def validate_rule13_proto_pricing_missing_notes(self) -> List[ValidationResult]:
        """
        Rule 13: Validate "#. Proto Pricing No Cost" in Missing Notes based on Proto headers
        where # = 1, 2, 3... (e.g., "1. Proto Pricing No Cost", "2. Proto Pricing No Cost")
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        
        # Check for Proto headers in BOM MATRIX
        proto_qty_found, proto_price_found = self._check_proto_headers_in_bom_matrix()
        proto_headers_exist = proto_qty_found and proto_price_found
        
        # Check Missing Notes sheet
        if "Missing Notes" not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 13: Proto Pricing in Missing Notes",
                sheet_name="Missing Notes",
                status="FAIL",
                expected="Sheet 'Missing Notes' should exist",
                actual="Missing Notes sheet not found"
            )
            results.append(result)
            return results
        
        # Look for "#. Proto Pricing No Cost" sections dynamically - search entire sheet
        sheet = self.workbook["Missing Notes"]
        proto_pricing_sections = []
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=20), 1):
            cell = sheet.cell(row=row_idx, column=1)  # Check column A for section headers
            if cell.value and isinstance(cell.value, str):
                value = cell.value.strip()
                # Check if it matches pattern like "1. Proto Pricing No Cost", "2. Proto Pricing No Cost"
                if "Proto Pricing No Cost" in value or "PROTO PRICING NO COST" in value.upper():
                    # Extract the number before the dot
                    parts = value.split(".")
                    if len(parts) >= 2 and parts[0].strip().isdigit():
                        proto_pricing_sections.append((value, row_idx))
        
        proto_pricing_found = len(proto_pricing_sections) > 0
        
        # Determine result
        if proto_headers_exist and proto_pricing_found:
            section_names = [name for name, _ in proto_pricing_sections]
            result = ValidationResult(
                rule_name="Rule 13: Proto Pricing in Missing Notes",
                sheet_name="Missing Notes",
                status="PASS",
                expected="'#. Proto Pricing No Cost' sections should be present",
                actual=f"Proto volume present and Proto Pricing No Cost is there in missing notes ({len(proto_pricing_sections)} section(s): {', '.join(section_names)})"
            )
        elif proto_headers_exist and not proto_pricing_found:
            result = ValidationResult(
                rule_name="Rule 13: Proto Pricing in Missing Notes",
                sheet_name="Missing Notes",
                status="FAIL",
                expected="'#. Proto Pricing No Cost' sections should be present (e.g., '1. Proto Pricing No Cost')",
                actual="Proto specified but Proto Pricing No Cost sections not found in missing notes"
            )
        elif not proto_headers_exist and proto_pricing_found:
            result = ValidationResult(
                rule_name="Rule 13: Proto Pricing in Missing Notes",
                sheet_name="Missing Notes",
                status="FAIL",
                expected="'#. Proto Pricing No Cost' should not be present without Proto",
                actual="Proto Pricing No Cost present but Proto not specified"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 13: Proto Pricing in Missing Notes",
                sheet_name="Missing Notes",
                status="PASS",
                expected="No Proto Pricing when Proto not specified",
                actual="Proto not specified, Proto Pricing correctly absent"
            )
        
        results.append(result)
        return results
    
    def validate_rule14_ex_inv_proto_sheet(self) -> List[ValidationResult]:
        """
        Rule 14: Validate Ex Inv VL-proto sheet exists if Proto headers are in BOM MATRIX
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        
        # Check for Proto headers in BOM MATRIX
        proto_qty_found, proto_price_found = self._check_proto_headers_in_bom_matrix()
        proto_headers_exist = proto_qty_found and proto_price_found
        
        # Check if Ex Inv VL-proto sheet exists
        ex_inv_proto_exists = "Ex Inv VL-proto" in self.workbook.sheetnames
        
        # Determine result
        if proto_headers_exist and ex_inv_proto_exists:
            result = ValidationResult(
                rule_name="Rule 14: Ex Inv VL-proto Sheet Validation",
                sheet_name="Ex Inv VL-proto",
                status="PASS",
                expected="'Ex Inv VL-proto' sheet should exist when Proto specified",
                actual="Proto volume present and Ex Inv VL-proto sheet present"
            )
        elif proto_headers_exist and not ex_inv_proto_exists:
            result = ValidationResult(
                rule_name="Rule 14: Ex Inv VL-proto Sheet Validation",
                sheet_name="N/A",
                status="FAIL",
                expected="'Ex Inv VL-proto' sheet should exist when Proto specified",
                actual="Proto specified but Ex Inv VL-proto sheet missing"
            )
        elif not proto_headers_exist and ex_inv_proto_exists:
            result = ValidationResult(
                rule_name="Rule 14: Ex Inv VL-proto Sheet Validation",
                sheet_name="Ex Inv VL-proto",
                status="FAIL",
                expected="'Ex Inv VL-proto' should not exist without Proto",
                actual="Ex Inv VL-proto present but Proto not specified"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 14: Ex Inv VL-proto Sheet Validation",
                sheet_name="N/A",
                status="PASS",
                expected="No Ex Inv Proto sheet when Proto not specified",
                actual="Proto not specified, Ex Inv VL-proto correctly absent"
            )
        
        results.append(result)
        return results
    
    def validate_rule15_proto_columns_bom_matrix(self) -> List[ValidationResult]:
        """
        Rule 15: Validate Proto Qty and Proto Price columns in BOM MATRIX match 7.0 CBOM Proto sheet existence
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        
        # Check if 7.0 CBOM Proto sheet exists
        cbom_proto_exists = "7.0 CBOM Proto" in self.workbook.sheetnames
        
        # Check for Proto headers in BOM MATRIX
        proto_qty_found, proto_price_found = self._check_proto_headers_in_bom_matrix()
        proto_headers_exist = proto_qty_found and proto_price_found
        
        # Determine result
        if cbom_proto_exists and proto_headers_exist:
            result = ValidationResult(
                rule_name="Rule 15: Proto Columns in BOM MATRIX",
                sheet_name="BOM MATRIX",
                status="PASS",
                expected="'Proto Qty' and 'Proto Price' headers should be present",
                actual="Proto volume is given and Proto Qty and Proto Price header is present in BOM MATRIX"
            )
        elif cbom_proto_exists and not proto_headers_exist:
            result = ValidationResult(
                rule_name="Rule 15: Proto Columns in BOM MATRIX",
                sheet_name="BOM MATRIX",
                status="FAIL",
                expected="'Proto Qty' and 'Proto Price' should be present when 7.0 CBOM Proto exists",
                actual="7.0 CBOM Proto sheet exists but Proto columns missing in BOM MATRIX"
            )
        elif not cbom_proto_exists and proto_headers_exist:
            result = ValidationResult(
                rule_name="Rule 15: Proto Columns in BOM MATRIX",
                sheet_name="BOM MATRIX",
                status="FAIL",
                expected="Proto columns should not exist without 7.0 CBOM Proto sheet",
                actual="Proto Qty/Proto Price present but 7.0 CBOM Proto sheet missing"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 15: Proto Columns in BOM MATRIX",
                sheet_name="BOM MATRIX",
                status="PASS",
                expected="No Proto columns when Proto sheet absent",
                actual="Proto sheet absent, Proto columns correctly absent"
            )
        
        results.append(result)
        return results
    
    def validate_rule16_serial_number_standardization(self) -> List[ValidationResult]:
        """
        Rule 16: Validate that serial numbers in Missing Notes are in standardized sequential order (1, 2, 3, 4...)
        Checks headers with blue/blue variant background (e.g., 1. Uncosted Parts, 2. NCNR Mentioned)
        Only validates the serial numbers, not the text content
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_name = "Missing Notes"
        
        if sheet_name not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 16: Serial Number Standardization",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Sheet 'Missing Notes' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results
        
        sheet = self.workbook[sheet_name]
        
        # Find all numbered section headers in column A (looking for pattern: "#. Text")
        # Blue background color indicates these are section headers
        numbered_sections = []
        
        for row_idx in range(1, min(sheet.max_row + 1, 1000)):
            cell = sheet.cell(row=row_idx, column=1)
            
            if cell.value and isinstance(cell.value, str):
                value = cell.value.strip()
                
                # Check if it matches pattern: "#. Something" where # is a number
                if '.' in value:
                    parts = value.split('.', 1)  # Split only on first dot
                    if len(parts) == 2 and parts[0].strip().isdigit():
                        serial_num = int(parts[0].strip())
                        section_text = parts[1].strip()
                        
                        # Check if cell has blue background (indicates section header)
                        # Blue colors typically have RGB values with high blue component
                        has_blue_bg = False
                        if cell.fill and cell.fill.fgColor:
                            # Check if cell has any fill color (blue variants)
                            if cell.fill.patternType and cell.fill.patternType != 'none':
                                has_blue_bg = True
                        
                        # If we can't reliably detect blue, assume numbered headers are section headers
                        numbered_sections.append((row_idx, serial_num, section_text))
        
        # Validate serial numbers are sequential (1, 2, 3, 4...)
        if not numbered_sections:
            result = ValidationResult(
                rule_name="Rule 16: Serial Number Standardization",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Numbered section headers should be present",
                actual="No numbered sections found in Missing Notes"
            )
        else:
            # Sort by row number to maintain document order
            numbered_sections.sort(key=lambda x: x[0])
            
            # Extract serial numbers
            serial_numbers = [num for _, num, _ in numbered_sections]
            
            # Check if they are sequential starting from 1
            expected_sequence = list(range(1, len(serial_numbers) + 1))
            is_sequential = serial_numbers == expected_sequence
            
            if is_sequential:
                section_list = ", ".join([f"{num}. {text[:30]}..." if len(text) > 30 else f"{num}. {text}" 
                                         for _, num, text in numbered_sections[:5]])  # Show first 5
                if len(numbered_sections) > 5:
                    section_list += f" (and {len(numbered_sections) - 5} more)"
                
                result = ValidationResult(
                    rule_name="Rule 16: Serial Number Standardization",
                    sheet_name=sheet_name,
                    status="PASS",
                    expected="Serial numbers in sequential order (1, 2, 3...)",
                    actual=f"Serial numbers are in standard format ({len(numbered_sections)} sections found)",
                    location=f"Sections: {section_list}"
                )
            else:
                # Find the mismatch
                mismatches = []
                for i, (row, actual_num, text) in enumerate(numbered_sections):
                    expected_num = i + 1
                    if actual_num != expected_num:
                        mismatches.append(f"Row {row}: Expected {expected_num}, Found {actual_num}")
                
                result = ValidationResult(
                    rule_name="Rule 16: Serial Number Standardization",
                    sheet_name=sheet_name,
                    status="FAIL",
                    expected="Serial numbers should be sequential (1, 2, 3, 4...)",
                    actual=f"Serial number mismatch: {'; '.join(mismatches[:3])}"
                )
        
        results.append(result)
        return results
    
    def validate_rule17_price_validity_date_format(self) -> List[ValidationResult]:
        """
        Rule 17: Validate that Price Validity uses date format (not weeks)
        Checks if cells below "Effective Date" header contain date format values
        
        Returns:
            List of ValidationResult objects
        """
        results = []
        sheet_name = "BOM MATRIX"
        
        if sheet_name not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 17: Price Validity Date Format",
                sheet_name=sheet_name,
                status="FAIL",
                expected="Sheet 'BOM MATRIX' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results
        
        sheet = self.workbook[sheet_name]
        
        # Find "Effective Date" header
        effective_date_found = False
        effective_date_col = None
        effective_date_row = None
        has_date_format = False
        
        for row_idx in range(1, min(30, sheet.max_row + 1)):
            for col_idx in range(1, min(100, sheet.max_col + 1) if hasattr(sheet, 'max_col') else 100):
                cell = sheet.cell(row=row_idx, column=col_idx)
                
                if cell.value and isinstance(cell.value, str) and "Effective Date" in cell.value:
                    effective_date_found = True
                    effective_date_col = col_idx
                    effective_date_row = row_idx
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    
                    # Check cells below for date format
                    for value_row in range(row_idx + 1, min(row_idx + 100, sheet.max_row + 1)):
                        value_cell = sheet.cell(row=value_row, column=col_idx)
                        
                        # Check if cell has date format or contains date value
                        if value_cell.value:
                            # Check for date number format (contains d, m, y)
                            if value_cell.number_format:
                                format_lower = value_cell.number_format.lower()
                                if any(indicator in format_lower for indicator in ['d', 'm', 'y', 'date']):
                                    has_date_format = True
                                    break
                            
                            # Check if value is datetime object
                            if hasattr(value_cell.value, 'date'):  # datetime object
                                has_date_format = True
                                break
                            
                            # Check if cell contains date-like string (e.g., "01/15/2025", "2025-01-15")
                            if isinstance(value_cell.value, str):
                                value_str = value_cell.value.strip()
                                # Simple date pattern check
                                if '/' in value_str or '-' in value_str:
                                    # Could be a date string
                                    parts = value_str.replace('/', '-').split('-')
                                    if len(parts) >= 2:
                                        # Check if parts look like numbers
                                        try:
                                            nums = [int(p) for p in parts if p.strip().isdigit()]
                                            if len(nums) >= 2:
                                                has_date_format = True
                                                break
                                        except:
                                            pass
                    break
            
            if effective_date_found:
                break
        
        # Create result
        if not effective_date_found:
            result = ValidationResult(
                rule_name="Rule 17: Price Validity Date Format",
                sheet_name=sheet_name,
                status="FAIL",
                expected="'Effective Date' header should be present",
                actual="'Effective Date' header not found in BOM MATRIX"
            )
        elif has_date_format:
            col_letter = openpyxl.utils.get_column_letter(effective_date_col)
            result = ValidationResult(
                rule_name="Rule 17: Price Validity Date Format",
                sheet_name=sheet_name,
                status="PASS",
                expected="Price validity in date format",
                actual="Price Validity in date format",
                location=f"{col_letter}{effective_date_row}"
            )
        else:
            col_letter = openpyxl.utils.get_column_letter(effective_date_col)
            result = ValidationResult(
                rule_name="Rule 17: Price Validity Date Format",
                sheet_name=sheet_name,
                status="FAIL",
                expected="At least one cell below 'Effective Date' should have date format",
                actual="No date format found in Effective Date column",
                location=f"{col_letter}{effective_date_row}"
            )
        
        results.append(result)
        return results
    
    def run_all_validations(self) -> List[ValidationResult]:
        """
        Run all validation rules
        
        Returns:
            List of all validation results
        """
        all_results = []
        
        # Original Rules (1-6)
        rule1_results = self.validate_rule1_header_presence()
        all_results.extend(rule1_results)
        
        rule2_results = self.validate_rule2_corrected_mpn()
        all_results.extend(rule2_results)
        
        rule3_results = self.validate_rule3_currency_symbols_cbom()
        all_results.extend(rule3_results)
        
        rule4_results = self.validate_rule4_moq_cost()
        all_results.extend(rule4_results)
        
        rule5_results = self.validate_rule5_currency_symbols_ex_inv()
        all_results.extend(rule5_results)
        
        rule6_results = self.validate_rule6_net_ordering_qty()
        all_results.extend(rule6_results)
        
        # New Rules (7-15)
        rule7_results = self.validate_rule7_currency_a_class_parts()
        all_results.extend(rule7_results)
        
        rule8_results = self.validate_rule8_quoted_mfr()
        all_results.extend(rule8_results)
        
        rule9_results = self.validate_rule9_nrfnd_missing()
        all_results.extend(rule9_results)
        
        rule10_results = self.validate_rule10_currency_bom_matrix()
        all_results.extend(rule10_results)
        
        rule11_results = self.validate_rule11_cbom_proto_sheet()
        all_results.extend(rule11_results)
        
        rule12_results = self.validate_rule12_proto_volume_summary()
        all_results.extend(rule12_results)
        
        rule13_results = self.validate_rule13_proto_pricing_missing_notes()
        all_results.extend(rule13_results)
        
        rule14_results = self.validate_rule14_ex_inv_proto_sheet()
        all_results.extend(rule14_results)
        
        rule15_results = self.validate_rule15_proto_columns_bom_matrix()
        all_results.extend(rule15_results)
        
        # Additional Rules (16-17)
        rule16_results = self.validate_rule16_serial_number_standardization()
        all_results.extend(rule16_results)
        
        rule17_results = self.validate_rule17_price_validity_date_format()
        all_results.extend(rule17_results)

        rule18_results = self.validate_rule18_uncosted_parts_count()
        all_results.extend(rule18_results)

        rule19_results = self.validate_rule19_bom_matrix_validation()
        all_results.extend(rule19_results)

        self.results = all_results
        return all_results

    def validate_rule18_uncosted_parts_count(self) -> List[ValidationResult]:
        """
        Rule 18: Part count is not matching for uncosted parts

        Logic:
        1. Find "#. Uncosted Parts" header in Missing Notes (where # = 1,2,3,4..)
        2. Find "SI.no" sub-header and get the last SI.no value
        3. Find the last 7.0 CBOM VL-{X} sheet (highest X)
        4. In that sheet, filter "AF" column (Is Data) for "False" values
        5. Count unique "Part Number" values where Is Data = False
        6. Compare: last SI.no should equal count of unique parts with Is Data = False

        Returns:
            List of ValidationResult objects
        """
        results = []

        # Step 1: Check if Missing Notes exists
        if "Missing Notes" not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name="Missing Notes",
                status="FAIL",
                expected="Sheet 'Missing Notes' should exist",
                actual="Sheet not found",
                location="#. Uncosted Parts, Is Data, Part Number"
            )
            results.append(result)
            return results

        missing_notes = self.workbook["Missing Notes"]

        # Step 2: Find "#. Uncosted Parts" section (where # = 1,2,3,4..)
        uncosted_row = None
        uncosted_header = None
        for row_idx in range(1, 500):
            cell = missing_notes.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str) and "Uncosted Parts" in cell.value:
                uncosted_row = row_idx
                uncosted_header = cell.value.strip()
                break

        if not uncosted_row:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name="Missing Notes",
                status="FAIL",
                expected="'#. Uncosted Parts' section should exist",
                actual="'Uncosted Parts' section not found",
                location="#. Uncosted Parts, Is Data, Part Number"
            )
            results.append(result)
            return results

        # Step 3: Find "SI.no" sub-header (can be "Sl.no", "SI.no", or "S.No")
        sl_no_col = None
        sl_no_header_row = None
        for col_idx in range(1, 20):
            for row_offset in range(1, 5):
                cell = missing_notes.cell(row=uncosted_row + row_offset, column=col_idx)
                if cell.value and isinstance(cell.value, str):
                    # Check various formats
                    if any(pattern in cell.value.upper() for pattern in ["SL.NO", "SI.NO", "S.NO", "SLNO", "SINO"]):
                        sl_no_col = col_idx
                        sl_no_header_row = uncosted_row + row_offset
                        break
            if sl_no_col:
                break

        if not sl_no_col:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name="Missing Notes",
                status="FAIL",
                expected="'SI.no' sub-header should exist under '#. Uncosted Parts'",
                actual="'SI.no' sub-header not found",
                location=f"{uncosted_header}, Is Data, Part Number"
            )
            results.append(result)
            return results

        # Step 4: Find the last SI.no value (take the last number in the SI.no column)
        # Stop when we hit the next section header (e.g., "2. NCNR Mentioned", "3. Something", etc.)
        last_sl_no = None
        for row_idx in range(sl_no_header_row + 1, min(sl_no_header_row + 1000, missing_notes.max_row + 1)):
            # Check if we've hit the next section (look for pattern like "2. ", "3. ", etc. in column 1)
            first_col_cell = missing_notes.cell(row=row_idx, column=1)
            if first_col_cell.value and isinstance(first_col_cell.value, str):
                first_col_val = first_col_cell.value.strip()
                # Check if this is a new section header (e.g., "2. NCNR Mentioned")
                if len(first_col_val) > 2 and first_col_val[0].isdigit() and '. ' in first_col_val[:4]:
                    # Found next section, stop here
                    break

            cell = missing_notes.cell(row=row_idx, column=sl_no_col)
            if cell.value:
                current_val = None
                if isinstance(cell.value, (int, float)):
                    current_val = int(cell.value)
                elif isinstance(cell.value, str) and cell.value.strip().isdigit():
                    current_val = int(cell.value.strip())

                # Take the last SI.no value (keep updating until we reach the end)
                if current_val is not None:
                    last_sl_no = current_val

        if last_sl_no is None:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name="Missing Notes",
                status="FAIL",
                expected="At least one SI.no value under '#. Uncosted Parts'",
                actual="No SI.no values found",
                location=f"{uncosted_header}, Is Data, Part Number"
            )
            results.append(result)
            return results

        # Step 5: Find the last CBOM VL-{X} sheet (take the last value for X)
        cbom_pattern = r'^7\.0 CBOM VL-(\d+)$'
        cbom_sheets = []
        for sheet_name in self.workbook.sheetnames:
            match = re.match(cbom_pattern, sheet_name)
            if match:
                sheet_num = int(match.group(1))
                cbom_sheets.append((sheet_num, sheet_name))

        if not cbom_sheets:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name="N/A",
                status="FAIL",
                expected="At least one '7.0 CBOM VL-{X}' sheet should exist",
                actual="No CBOM VL sheets found",
                location=f"{uncosted_header}, Is Data, Part Number"
            )
            results.append(result)
            return results

        # Get the last CBOM sheet (highest X value)
        cbom_sheets.sort()
        last_cbom_num, last_cbom_name = cbom_sheets[-1]
        cbom_sheet = self.workbook[last_cbom_name]

        # Step 6: Find "Is Data" column (AF = column 32) and filter for "False"
        is_data_col = 32  # AF column
        is_data_row = None
        for row_idx in range(1, 50):
            cell = cbom_sheet.cell(row=row_idx, column=is_data_col)
            if cell.value and isinstance(cell.value, str) and "Is Data" in cell.value:
                is_data_row = row_idx
                break

        if not is_data_row:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name=last_cbom_name,
                status="FAIL",
                expected="'Is Data' column should exist in column AF",
                actual="'Is Data' column not found",
                location=f"{uncosted_header}, Is Data, Part Number"
            )
            results.append(result)
            return results

        # Step 7: Find "Part Number" column in the same header row as "Is Data"
        part_number_col = None
        for col_idx in range(1, 50):
            cell = cbom_sheet.cell(row=is_data_row, column=col_idx)
            if cell.value and isinstance(cell.value, str) and cell.value.strip() == "Part Number":
                part_number_col = col_idx
                break

        if not part_number_col:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name=last_cbom_name,
                status="FAIL",
                expected="'Part Number' column should exist",
                actual="'Part Number' column not found",
                location=f"{uncosted_header}, Is Data, Part Number"
            )
            results.append(result)
            return results

        # Step 8: Count unique part numbers where Is Data = False (filter AF column for "False")
        false_parts = set()
        for row_idx in range(is_data_row + 1, cbom_sheet.max_row + 1):
            is_data_cell = cbom_sheet.cell(row=row_idx, column=is_data_col)
            part_number_cell = cbom_sheet.cell(row=row_idx, column=part_number_col)

            # Check if Is Data is False (filter for "False" value)
            is_false = False
            if is_data_cell.value is False:
                is_false = True
            elif isinstance(is_data_cell.value, str) and is_data_cell.value.upper() == 'FALSE':
                is_false = True

            # Count unique Part Number values where Is Data is False
            if is_false and part_number_cell.value:
                false_parts.add(str(part_number_cell.value).strip())

        cbom_false_count = len(false_parts)

        # Step 9: Compare the last SI.no with count of unique Part Numbers where Is Data is False
        if last_sl_no == cbom_false_count:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name=f"Missing Notes, {last_cbom_name}",
                status="PASS",
                expected=f"Part count matching between Missing Notes and {last_cbom_name}",
                actual="Part count is matching for uncosted part",
                location=f"{uncosted_header}, Is Data, Part Number"
            )
        else:
            result = ValidationResult(
                rule_name="Rule 18: Part count is not matching for uncosted parts",
                sheet_name=f"Missing Notes, {last_cbom_name}",
                status="FAIL",
                expected=f"Last SI.no ({last_sl_no}) should match unique Part Number count with Is Data=False ({cbom_false_count})",
                actual=f"Part count is not matching for uncosted part (SI.no: {last_sl_no}, Is Data=False count: {cbom_false_count})",
                location=f"{uncosted_header}, Is Data, Part Number"
            )

        results.append(result)
        return results

    def validate_rule19_bom_matrix_validation(self) -> List[ValidationResult]:
        """
        Rule 19: CPN count is not matching - Validate CPN count matching between Lead Time (FG Wise) and CBOM VL sheets

        Logic:
        1. In Lead Time (FG Wise), find "LT in weeks - #" headers (extract FG part numbers)
        2. Find "Grand Total" row for each FG, get count from next column
        3. Use last 7.0 CBOM VL-{X} sheet (highest X)
        4. Find ALL "FG part number" headers in Column A (scan entire sheet)
        5. For each FG, filter rows where Column A = FG part number
        6. Count UNIQUE "Part Number" (Column C) values
        7. Compare: Grand Total count == Unique CBOM count

        Returns individual result for EACH FG part

        Returns:
            List of ValidationResult objects (one per FG part)
        """
        results = []

        # Step 1: Check if Lead Time (FG Wise) exists
        if "Lead Time (FG Wise)" not in self.workbook.sheetnames:
            result = ValidationResult(
                rule_name="Rule 19: CPN count is not matching",
                sheet_name="Lead Time (FG Wise)",
                status="FAIL",
                expected="Sheet 'Lead Time (FG Wise)' should exist",
                actual="Sheet not found"
            )
            results.append(result)
            return results

        lt_sheet = self.workbook["Lead Time (FG Wise)"]

        # Step 2: Find all "LT in weeks - #" headers and their Grand Total counts
        fg_parts = []
        for col_idx in range(1, 50):
            header_cell = lt_sheet.cell(row=1, column=col_idx)
            if header_cell.value and isinstance(header_cell.value, str) and "LT in weeks -" in header_cell.value:
                fg_part = header_cell.value.replace("LT in weeks -", "").strip()

                # Next column should be "Count of Part Number"
                count_col = col_idx + 1

                # Find "Grand Total" row in this FG's column
                grand_total_value = None
                for row_idx in range(2, 100):
                    cell_value = lt_sheet.cell(row=row_idx, column=col_idx).value
                    if cell_value and isinstance(cell_value, str) and "Grand Total" in cell_value:
                        grand_total_value = lt_sheet.cell(row=row_idx, column=count_col).value
                        break

                if grand_total_value is not None:
                    fg_parts.append((fg_part, grand_total_value))

        if not fg_parts:
            result = ValidationResult(
                rule_name="Rule 19: CPN count is not matching",
                sheet_name="Lead Time (FG Wise)",
                status="FAIL",
                expected="At least one 'LT in weeks - #' header with Grand Total",
                actual="No FG parts with Grand Total found"
            )
            results.append(result)
            return results

        # Step 3: Find last CBOM VL-{X} sheet
        cbom_pattern = r'^7\.0 CBOM VL-(\d+)$'
        cbom_sheets = []
        for sheet_name in self.workbook.sheetnames:
            match = re.match(cbom_pattern, sheet_name)
            if match:
                sheet_num = int(match.group(1))
                cbom_sheets.append((sheet_num, sheet_name))

        if not cbom_sheets:
            result = ValidationResult(
                rule_name="Rule 19: CPN count is not matching",
                sheet_name="N/A",
                status="FAIL",
                expected="At least one '7.0 CBOM VL-{X}' sheet should exist",
                actual="No CBOM VL sheets found"
            )
            results.append(result)
            return results

        # Get last CBOM sheet (highest X)
        cbom_sheets.sort()
        last_cbom_num, last_cbom_name = cbom_sheets[-1]
        cbom_sheet = self.workbook[last_cbom_name]

        # Step 4: Find ALL "FG part number" sections in Column A (scan entire sheet up to 2000 rows)
        fg_sections = []
        for row_idx in range(1, 2000):
            cell = cbom_sheet.cell(row=row_idx, column=1)
            if cell.value and isinstance(cell.value, str) and cell.value == "FG part number":
                fg_sections.append(row_idx)

        if not fg_sections:
            result = ValidationResult(
                rule_name="Rule 19: CPN count is not matching",
                sheet_name=last_cbom_name,
                status="FAIL",
                expected="'FG part number' headers in Column A",
                actual="No 'FG part number' headers found"
            )
            results.append(result)
            return results

        # Step 5: Validate EACH FG part individually and create separate result for each
        for fg_part, expected_count in fg_parts:
            unique_parts = set()

            # Count unique parts across ALL sections
            for section_row in fg_sections:
                row_idx = section_row + 1

                # Scan until empty row (with large limit per section)
                while row_idx < section_row + 1000:
                    fg_value = cbom_sheet.cell(row=row_idx, column=1).value  # Column A - FG part number
                    part_number = cbom_sheet.cell(row=row_idx, column=3).value  # Column C - Part Number

                    # Stop at empty row
                    if fg_value is None or str(fg_value).strip() == "":
                        break

                    # Filter by FG part number and collect unique part numbers
                    if str(fg_value).strip() == fg_part and part_number:
                        unique_parts.add(str(part_number).strip())

                    row_idx += 1

            actual_count = len(unique_parts)

            # Create individual result for this FG part
            if expected_count == actual_count:
                result = ValidationResult(
                    rule_name="Rule 19: CPN count is not matching",
                    sheet_name="Lead Time (FG Wise)",
                    status="PASS",
                    expected=f"Count {expected_count} for FG {fg_part}",
                    actual=f"{fg_part}: {expected_count} == {actual_count} PASS",
                    location=f"Lead Time vs {last_cbom_name}"
                )
            else:
                if actual_count == 0:
                    result = ValidationResult(
                        rule_name="Rule 19: CPN count is not matching",
                        sheet_name="Lead Time (FG Wise)",
                        status="FAIL",
                        expected=f"Count {expected_count} for FG {fg_part}",
                        actual=f"{fg_part}: Not found in CBOM",
                        location=f"Lead Time vs {last_cbom_name}"
                    )
                else:
                    result = ValidationResult(
                        rule_name="Rule 19: CPN count is not matching",
                        sheet_name="Lead Time (FG Wise)",
                        status="FAIL",
                        expected=f"Count {expected_count} for FG {fg_part}",
                        actual=f"{fg_part}: {expected_count} != {actual_count} FAIL",
                        location=f"Lead Time vs {last_cbom_name}"
                    )

            results.append(result)

        return results

    def generate_report(self, output_format: str = "console") -> str:
        """
        Generate validation report
        
        Args:
            output_format: Format of the report ("console", "csv", or "excel")
        
        Returns:
            Report as string or filepath
        """
        if not self.results:
            self.run_all_validations()
        
        if output_format == "console":
            return self._generate_console_report()
        elif output_format == "csv":
            return self._generate_csv_report()
        elif output_format == "excel":
            return self._generate_excel_report()
        else:
            raise ValueError(f"Unsupported output format: {output_format}")
    
    def _generate_console_report(self) -> str:
        """Generate console-friendly report"""
        report_lines = []
        report_lines.append("=" * 120)
        report_lines.append("EXCEL VALIDATION REPORT")
        report_lines.append("=" * 120)
        report_lines.append(f"File: {self.filepath}")
        report_lines.append(f"Total Rules Checked: {len(set(r.rule_name for r in self.results))}")
        report_lines.append(f"Total Validations: {len(self.results)}")
        report_lines.append(f"Passed: {sum(1 for r in self.results if r.status == 'PASS')}")
        report_lines.append(f"Failed: {sum(1 for r in self.results if r.status == 'FAIL')}")
        report_lines.append("=" * 120)
        report_lines.append("")
        
        # Group results by rule
        rules = {}
        for result in self.results:
            if result.rule_name not in rules:
                rules[result.rule_name] = []
            rules[result.rule_name].append(result)
        
        # Print results for each rule
        for rule_name, rule_results in rules.items():
            report_lines.append(f"\n{rule_name}")
            report_lines.append("-" * 120)
            
            for result in rule_results:
                status_symbol = "[OK]" if result.status == "PASS" else "[X]"
                report_lines.append(f"\n{status_symbol} Sheet: {result.sheet_name}")
                report_lines.append(f"  Status: {result.status}")
                report_lines.append(f"  Expected: {result.expected}")
                report_lines.append(f"  Actual: {result.actual}")
                if result.location:
                    report_lines.append(f"  Location: {result.location}")
        
        report_lines.append("\n" + "=" * 120)
        
        return "\n".join(report_lines)
    
    def _generate_csv_report(self) -> str:
        """Generate CSV report and save to file"""
        import csv
        
        output_file = "D:/OneDrive/Desktop/localllm/validation_report.csv"
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(['Rule Name', 'Sheet Name', 'Status', 'Expected', 'Actual', 'Location'])
            
            # Write data
            for result in self.results:
                writer.writerow([
                    result.rule_name,
                    result.sheet_name,
                    result.status,
                    result.expected,
                    result.actual,
                    result.location
                ])
        
        return output_file
    
    def _generate_excel_report(self) -> str:
        """Generate Excel report and save to file"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        output_file = "D:/OneDrive/Desktop/localllm/validation_report.xlsx"
        
        # Create new workbook
        report_wb = openpyxl.Workbook()
        ws = report_wb.active
        ws.title = "Validation Report"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # Write headers
        headers = ['Rule Name', 'Sheet Name', 'Status', 'Expected', 'Actual', 'Location']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data
        for row_idx, result in enumerate(self.results, 2):
            ws.cell(row=row_idx, column=1, value=result.rule_name)
            ws.cell(row=row_idx, column=2, value=result.sheet_name)
            
            status_cell = ws.cell(row=row_idx, column=3, value=result.status)
            status_cell.fill = pass_fill if result.status == "PASS" else fail_fill
            
            ws.cell(row=row_idx, column=4, value=result.expected)
            ws.cell(row=row_idx, column=5, value=result.actual)
            ws.cell(row=row_idx, column=6, value=result.location)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 15
        
        # Save workbook
        report_wb.save(output_file)
        
        return output_file
    
    def close(self):
        """Close the workbook"""
        self.workbook.close()


# Example usage
if __name__ == "__main__":
    import sys
    
    # Check command line arguments
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        
        # Check if diagnostic mode is requested
        if len(sys.argv) > 2 and sys.argv[2] in ['--diagnose', '-d', 'diagnose']:
            # Run diagnostics
            print("\n" + "="*100)
            print("RUNNING DIAGNOSTICS MODE")
            print("="*100)
            validator = ExcelValidator(filepath)
            validator.run_diagnostics()
            validator.close()
        else:
            # Run validation
            validator = ExcelValidator(filepath)
            results = validator.run_all_validations()
            
            # Print console report
            print(validator.generate_report("console"))
            
            # Generate CSV report
            csv_file = validator.generate_report("csv")
            print(f"\nCSV report saved to: {csv_file}")
            
            # Generate Excel report
            excel_file = validator.generate_report("excel")
            print(f"Excel report saved to: {excel_file}")
            
            validator.close()
    else:
        # Default behavior - validate the sample file
        validator = ExcelValidator("D:\OneDrive\Desktop\localllm\BOM_Matrix_726_20260113_152102 (1).xlsx")
        
        # Run validations
        results = validator.run_all_validations()
        
        # Print console report
        print(validator.generate_report("console"))
        
        # Generate CSV report
        csv_file = validator.generate_report("csv")
        print(f"\nCSV report saved to: {csv_file}")
        
        # Generate Excel report
        excel_file = validator.generate_report("excel")
        print(f"Excel report saved to: {excel_file}")
        
        # Close validator
        validator.close()
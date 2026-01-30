"""
Excel Validation Tool for Quote Win Files in CAM
Template is embedded - users only need to provide input file.
Updated with pattern-based header validation.
"""

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
import sys
import re
from datetime import datetime


# EMBEDDED TEMPLATE CONFIGURATION
TEMPLATE_CONFIG = {
    "summary_header_row": 12,  # Row with Group By Field and Ext Vol headers
    "header_row": 17,           # Row with Project, Part Number, etc.
    "data_start_row": 18,
    "project_row": 3,
    "project_value_column": 4,
    
    # Required static headers in summary row (Row 12)
    "required_summary_static_headers": [
        "Group By Field"
    ],
    
    # Required dynamic patterns in summary row (Row 12)
    "required_summary_dynamic_patterns": [
        r"^Ext Vol \(Splits\) #\d+$",
        r"^% Ext Vol Qty #\d+$",
        r"^Ext Part Vol Cost \(Splits\) #\d+ \(Conv\.\)$",
        r"^% Ext Vol Cost \(Splits\) #\d+$"
    ],
    
    # Required static headers in main header row (Row 16)
    "required_static_headers": [
        "Project",
        "Part Number",
        "Part Description",
        "Commodity",
        "Mfg Name",
        "Mfg Part Number",
        "Currency (Original)",
        "Supp Name",
        "Pkg Qty",
        "MOQ",
        "Lead Time",
        "Part Qty",
        "Corrected MPN",
        "Long Comment",
        "Price Type",
        "No Bid Reason",
        "Short Comment",
        "NCNR",
        "RFQ Number",
        "Eff Date",
        "Exp Date",
        "Quote Validity",
        "Part Status"
    ],
    
    # Required dynamic header patterns in main header row (Row 16, where X = 1, 2, 3, ...)
    "required_dynamic_patterns": [
        r"^Cost #\d+ \(Conv\.\)$",
        r"^Price \(Original\) #\d+$",
        r"^Awarded Volume #\d+$",
        r"^Award #\d+$",
        r"^Source #\d+$"
    ]
}


class ExcelValidator:
    """Validates Quote Win Excel files against embedded template."""
    
    def __init__(self, input_path, output_path=None):
        """
        Initialize the validator.
        
        Args:
            input_path: Path to the input Excel file to validate
            output_path: Path to save the validated file (optional)
        """
        self.input_path = input_path
        self.output_path = output_path or input_path.replace('.xlsx', '_validated.xlsx')
        
        self.input_wb = None
        self.input_sheet = None
        
        # Template configuration
        self.summary_header_row = TEMPLATE_CONFIG["summary_header_row"]
        self.required_summary_static_headers = TEMPLATE_CONFIG["required_summary_static_headers"]
        self.required_summary_dynamic_patterns = TEMPLATE_CONFIG["required_summary_dynamic_patterns"]
        self.required_static_headers = TEMPLATE_CONFIG["required_static_headers"]
        self.required_dynamic_patterns = TEMPLATE_CONFIG["required_dynamic_patterns"]
        self.header_row = TEMPLATE_CONFIG["header_row"]
        self.data_start_row = TEMPLATE_CONFIG["data_start_row"]
        self.project_row = TEMPLATE_CONFIG["project_row"]
        self.project_value_col = TEMPLATE_CONFIG["project_value_column"]
        
        self.validation_results = {
            'header_validation': {'passed': True, 'message': '', 'missing_headers': [], 'invalid_headers': []},
            'project_name_validation': {'passed': True, 'message': ''},
            'filter_validation': {'passed': True, 'message': ''},
            'award_validation': {'passed': True, 'issues': [], 'award_column_results': {}}
        }
        
    def load_file(self):
        """Load the input Excel file."""
        try:
            self.input_wb = openpyxl.load_workbook(self.input_path)
            self.input_sheet = self.input_wb.active
            
            print("[OK] File loaded successfully")
            return True
        except Exception as e:
            print(f"[X] Error loading file: {str(e)}")
            return False
    
    def _matches_any_pattern(self, header):
        """Check if a header matches any of the required dynamic patterns."""
        if not header:
            return False
        for pattern in self.required_dynamic_patterns:
            if re.match(pattern, str(header).strip()):
                return True
        return False
    
    def validate_headers(self):
        """
        Validate that required headers in input file are present.
        Checks headers in both summary row (Row 12) and main header row (Row 16).
        Only checks if required headers exist - does not flag extra headers.
        """
        print("\n1. Validating Headers...")

        # Get headers from summary row (Row 12)
        summary_headers = []
        for cell in self.input_sheet[self.summary_header_row]:
            if cell.value:
                summary_headers.append(str(cell.value).strip())

        # Get headers from main header row (Row 16)
        main_headers = []
        for cell in self.input_sheet[self.header_row]:
            if cell.value:
                main_headers.append(str(cell.value).strip())

        print(f"  Found {len(summary_headers)} headers in summary row (Row {self.summary_header_row})")
        print(f"  Found {len(main_headers)} headers in main header row (Row {self.header_row})")

        # Track missing headers only (not invalid/extra headers)
        missing_summary_static = []
        missing_summary_dynamic_groups = {}
        missing_main_static = []
        missing_main_dynamic_groups = {}

        # === VALIDATE SUMMARY ROW (Row 12) ===

        # Check static headers in summary row
        for required_header in self.required_summary_static_headers:
            if required_header not in summary_headers:
                missing_summary_static.append(required_header)

        # Check dynamic headers in summary row - group by pattern type
        summary_dynamic_headers_found = {}
        for pattern in self.required_summary_dynamic_patterns:
            summary_dynamic_headers_found[pattern] = []

        for header in summary_headers:
            for pattern in self.required_summary_dynamic_patterns:
                if re.match(pattern, header):
                    summary_dynamic_headers_found[pattern].append(header)
                    break

        # Check that each dynamic pattern in summary row has at least one instance
        for pattern in self.required_summary_dynamic_patterns:
            if not summary_dynamic_headers_found[pattern]:
                pattern_name = pattern.replace(r"^", "").replace(r"$", "").replace(r"\d+", "X").replace(r"\(", "(").replace(r"\)", ")").replace(r"\.", ".")
                missing_summary_dynamic_groups[pattern_name] = pattern

        # === VALIDATE MAIN HEADER ROW (Row 16) ===

        # Check static headers in main row
        for required_header in self.required_static_headers:
            if required_header not in main_headers:
                missing_main_static.append(required_header)

        # Check dynamic headers in main row - group by pattern type
        main_dynamic_headers_found = {}
        for pattern in self.required_dynamic_patterns:
            main_dynamic_headers_found[pattern] = []

        for header in main_headers:
            for pattern in self.required_dynamic_patterns:
                if re.match(pattern, header):
                    main_dynamic_headers_found[pattern].append(header)
                    break

        # Check that each dynamic pattern in main row has at least one instance
        for pattern in self.required_dynamic_patterns:
            if not main_dynamic_headers_found[pattern]:
                pattern_name = pattern.replace(r"^", "").replace(r"$", "").replace(r"\d+", "X").replace(r"\(", "(").replace(r"\)", ")").replace(r"\.", ".")
                missing_main_dynamic_groups[pattern_name] = pattern

        # === DETERMINE VALIDATION RESULT ===
        # Only check if required headers are present (don't flag extra headers)
        validation_passed = (
            len(missing_summary_static) == 0 and
            len(missing_summary_dynamic_groups) == 0 and
            len(missing_main_static) == 0 and
            len(missing_main_dynamic_groups) == 0
        )
        
        if not validation_passed:
            self.validation_results['header_validation']['passed'] = False

            all_missing = (missing_summary_static + list(missing_summary_dynamic_groups.keys()) +
                          missing_main_static + list(missing_main_dynamic_groups.keys()))

            # Build message showing which headers are not matching
            missing_headers_str = ", ".join(all_missing)
            self.validation_results['header_validation']['message'] = f"Headers are not matching: {missing_headers_str}"
            self.validation_results['header_validation']['missing_headers'] = all_missing

            # Build detailed comment showing all missing headers
            comment_text = "Headers are not matching.\n\n"
            comment_text += "Missing Headers:\n"
            for header in all_missing:
                comment_text += f"  - {header}\n"

            if missing_summary_static or missing_summary_dynamic_groups:
                comment_text += f"\nSUMMARY ROW (Row {self.summary_header_row}):\n"
                if missing_summary_static:
                    for header in missing_summary_static:
                        comment_text += f"  - {header}\n"
                if missing_summary_dynamic_groups:
                    for pattern_name in missing_summary_dynamic_groups.keys():
                        comment_text += f"  - {pattern_name}\n"

            if missing_main_static or missing_main_dynamic_groups:
                comment_text += f"\nMAIN HEADER ROW (Row {self.header_row}):\n"
                if missing_main_static:
                    for header in missing_main_static:
                        comment_text += f"  - {header}\n"
                if missing_main_dynamic_groups:
                    for pattern_name in missing_main_dynamic_groups.keys():
                        comment_text += f"  - {pattern_name}\n"

            # Add comment to summary row first cell
            cell = self.input_sheet.cell(row=self.summary_header_row, column=1)
            comment = Comment(comment_text.strip(), "Validation Tool")
            cell.comment = comment

            # Highlight the header rows
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.fill = yellow_fill

            # Also highlight main header row
            cell_main = self.input_sheet.cell(row=self.header_row, column=1)
            cell_main.fill = yellow_fill

            print(f"  [X] Header validation failed - Headers are not matching")
            print(f"  Missing headers:")
            for header in all_missing:
                print(f"    - {header}")
            print(f"  Comment added to cell A{self.summary_header_row}")
        else:
            print(f"  [OK] All required headers present in both rows")
            
            # Show dynamic header counts for summary row
            print(f"\n  Summary Row (Row {self.summary_header_row}) Dynamic Headers:")
            for pattern in self.required_summary_dynamic_patterns:
                if summary_dynamic_headers_found[pattern]:
                    pattern_name = pattern.replace(r"^", "").replace(r"$", "").replace(r"\d+", "X").replace(r"\(", "(").replace(r"\)", ")").replace(r"\.", ".")
                    print(f"    - Found {len(summary_dynamic_headers_found[pattern])} {pattern_name} headers")
            
            # Show dynamic header counts for main row
            print(f"\n  Main Header Row (Row {self.header_row}) Dynamic Headers:")
            for pattern in self.required_dynamic_patterns:
                if main_dynamic_headers_found[pattern]:
                    pattern_name = pattern.replace(r"^", "").replace(r"$", "").replace(r"\d+", "X").replace(r"\(", "(").replace(r"\)", ")").replace(r"\.", ".")
                    print(f"    - Found {len(main_dynamic_headers_found[pattern])} {pattern_name} headers")
    
    def validate_project_name(self):
        """Validate that the Project value matches between the Project row and header section."""
        print("\n2. Validating Project Name...")
        
        # Get project value from project row
        project_cell = self.input_sheet.cell(row=self.project_row, column=self.project_value_col)
        project_value = project_cell.value
        
        # Get project value from data section (first non-empty value in Project column)
        header_project_value = None
        for row_num in range(self.data_start_row, self.input_sheet.max_row + 1):
            cell_value = self.input_sheet.cell(row=row_num, column=1).value
            if cell_value is not None:
                header_project_value = cell_value
                break
        
        print(f"  Project value in row {self.project_row}: {project_value}")
        print(f"  Project value in data section: {header_project_value}")
        
        # Compare values
        if project_value != header_project_value:
            self.validation_results['project_name_validation']['passed'] = False
            self.validation_results['project_name_validation']['message'] = "Project name not matching."
            
            # Add comment to project cell
            comment = Comment(
                f"Project name not matching.\nExpected: {header_project_value}\nGot: {project_value}",
                "Validation Tool"
            )
            project_cell.comment = comment
            
            # Highlight the cell
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            project_cell.fill = yellow_fill
            
            print(f"  [X] Project names do not match")
            print(f"  Comment added to cell D{self.project_row}")
        else:
            print(f"  [OK] Project names match")
    
    def validate_filters(self):
        """Check if any filters are applied to the Excel file."""
        print("\n3. Validating Filters...")

        # Check for auto filters
        if self.input_sheet.auto_filter and self.input_sheet.auto_filter.ref:
            filter_ref = self.input_sheet.auto_filter.ref
            self.validation_results['filter_validation']['passed'] = False
            self.validation_results['filter_validation']['message'] = f"Filter applied at {filter_ref} – please remove the filter."

            # Get the first cell of the filter range (where filter is applied)
            filter_start_cell = filter_ref.split(':')[0] if ':' in filter_ref else filter_ref

            # Add comment to the filter location
            cell = self.input_sheet[filter_start_cell]
            comment_text = f"Filter applied at {filter_ref} – please remove the filter."
            comment = Comment(comment_text, "Validation Tool")
            cell.comment = comment

            # Highlight the filter header cells
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.fill = yellow_fill

            print(f"  [X] Filter is applied")
            print(f"    Location: {filter_ref}")
            print(f"  Comment added to cell {filter_start_cell}")
        else:
            print(f"  [OK] No filters applied")
    
    def validate_awards(self):
        """
        Validate Award #X columns for each unique part number.

        Award validation logic (per Award column):
        - Each Award column (Award #1, Award #2, etc.) is validated separately
        - For each Award column, each unique part number must have exactly ONE row with Award = 100
        - If NO Award = 100 for a part number in that column → FAIL
        - If MORE THAN ONE row has Award = 100 for a part number in that column → FAIL
        """
        print("\n4. Validating Awards...")

        # Find Award #X columns
        award_columns = {}
        headers = list(self.input_sheet[self.header_row])

        for i, cell in enumerate(headers, 1):
            if cell.value and re.match(r'^Award #\d+$', str(cell.value).strip()):
                award_num = re.search(r'#(\d+)', str(cell.value)).group(1)
                award_columns[award_num] = i

        print(f"  Found {len(award_columns)} award columns: {list(award_columns.keys())}")

        if not award_columns:
            print(f"  [X] No Award columns found")
            return

        # Find Part Number column
        part_number_col = None
        for i, cell in enumerate(headers, 1):
            if cell.value and str(cell.value).strip() == 'Part Number':
                part_number_col = i
                break

        if not part_number_col:
            print(f"  [X] Part Number column not found")
            return

        print(f"  Part Number column: {part_number_col}")

        # Get all unique part numbers and their rows
        part_number_rows = {}  # {part_number: [row_nums]}
        for row_num in range(self.data_start_row, self.input_sheet.max_row + 1):
            part_number = self.input_sheet.cell(row=row_num, column=part_number_col).value
            # Skip None and empty part numbers
            if part_number is not None and str(part_number).strip() != '':
                part_number_str = str(part_number).strip()
                if part_number_str not in part_number_rows:
                    part_number_rows[part_number_str] = []
                part_number_rows[part_number_str].append(row_num)

        print(f"  Found {len(part_number_rows)} unique part numbers")

        # Initialize results structure for each award column
        self.validation_results['award_validation']['award_column_results'] = {}

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        total_issues = 0

        # Validate each Award column separately
        for award_num, col_num in sorted(award_columns.items(), key=lambda x: int(x[0])):
            print(f"\n  Validating Award #{award_num}...")

            no_award_parts = []
            multiple_award_parts = []

            # For each part number, check Award column
            for part_number, row_nums in part_number_rows.items():
                rows_with_award = []

                for row_num in row_nums:
                    award_value = self.input_sheet.cell(row=row_num, column=col_num).value

                    if award_value is not None:
                        is_100 = False
                        if isinstance(award_value, (int, float)) and award_value == 100:
                            is_100 = True
                        elif isinstance(award_value, str) and award_value.strip() == '100':
                            is_100 = True

                        if is_100:
                            rows_with_award.append(row_num)

                award_count = len(rows_with_award)

                if award_count == 0:
                    # No Award = 100 for this part number in this Award column
                    no_award_parts.append(part_number)  # part_number is already a string
                    issue_message = f"Award #{award_num} is not present for Part Number: {part_number}"
                    self.validation_results['award_validation']['issues'].append(issue_message)

                    # Add comment to the Part Number cell to indicate which part number is missing award
                    first_row = row_nums[0]
                    part_number_cell = self.input_sheet.cell(row=first_row, column=part_number_col)
                    comment = Comment(issue_message, "Validation Tool")
                    part_number_cell.comment = comment
                    part_number_cell.fill = yellow_fill

                elif award_count > 1:
                    # Multiple rows have Award = 100 for this part number in this Award column
                    multiple_award_parts.append(part_number)  # part_number is already a string
                    awarded_rows_str = ', '.join([str(r) for r in rows_with_award])
                    issue_message = f"Multiple Awards (100) found in Award #{award_num} for Part Number: {part_number} at rows: {awarded_rows_str}"
                    self.validation_results['award_validation']['issues'].append(issue_message)

                    # Add comment to all rows that have Award = 100 in this column
                    for row_num in rows_with_award:
                        cell = self.input_sheet.cell(row=row_num, column=col_num)
                        comment = Comment(issue_message, "Validation Tool")
                        cell.comment = comment
                        cell.fill = yellow_fill

            # Store results for this Award column
            self.validation_results['award_validation']['award_column_results'][award_num] = {
                'no_award_parts': no_award_parts,
                'multiple_award_parts': multiple_award_parts
            }

            column_issues = len(no_award_parts) + len(multiple_award_parts)
            total_issues += column_issues

            if column_issues > 0:
                print(f"    [X] Award #{award_num} validation failed")
                if no_award_parts:
                    print(f"      - Part numbers without Award = 100: {len(no_award_parts)}")
                if multiple_award_parts:
                    print(f"      - Part numbers with multiple Awards = 100: {len(multiple_award_parts)}")
            else:
                print(f"    [OK] Award #{award_num} validation passed")

        if total_issues > 0:
            self.validation_results['award_validation']['passed'] = False
            self.validation_results['award_validation']['message'] = f"Award validation failed - {total_issues} issues found"
            print(f"\n  [X] Award validation failed - {total_issues} total issues")
            print(f"  Comments added to respective cells")
        else:
            self.validation_results['award_validation']['message'] = "Award validation passed"
            print(f"\n  [OK] Award validation passed - All Award columns validated successfully")
    
    def save_results(self):
        """Save the validated Excel file with comments."""
        try:
            self.input_wb.save(self.output_path)
            print(f"\n[OK] Validation results saved to: {self.output_path}")
            return True
        except Exception as e:
            print(f"\n[X] Error saving file: {str(e)}")
            return False
    
    def print_summary(self):
        """Print validation summary."""
        print("\n" + "="*80)
        print("VALIDATION SUMMARY")
        print("="*80)
        
        all_passed = True
        
        # Header validation
        if self.validation_results['header_validation']['passed']:
            print("[OK] Header Validation: PASSED")
        else:
            print("[X] Header Validation: FAILED")
            print(f"  → {self.validation_results['header_validation']['message']}")
            if self.validation_results['header_validation']['missing_headers']:
                print(f"  → Missing headers: {len(self.validation_results['header_validation']['missing_headers'])}")
            if self.validation_results['header_validation']['invalid_headers']:
                print(f"  → Invalid headers: {len(self.validation_results['header_validation']['invalid_headers'])}")
            all_passed = False
        
        # Project name validation
        if self.validation_results['project_name_validation']['passed']:
            print("[OK] Project Name Validation: PASSED")
        else:
            print("[X] Project Name Validation: FAILED")
            print(f"  → {self.validation_results['project_name_validation']['message']}")
            all_passed = False
        
        # Filter validation
        if self.validation_results['filter_validation']['passed']:
            print("[OK] Filter Validation: PASSED")
        else:
            print("[X] Filter Validation: FAILED")
            print(f"  → {self.validation_results['filter_validation']['message']}")
            all_passed = False
        
        # Award validation
        if self.validation_results['award_validation']['passed']:
            print("[OK] Award Validation: PASSED")
        else:
            print("[X] Award Validation: FAILED")
            print(f"  → Found {len(self.validation_results['award_validation']['issues'])} issues")
            for issue in self.validation_results['award_validation']['issues'][:5]:  # Show first 5
                print(f"  → {issue}")
            if len(self.validation_results['award_validation']['issues']) > 5:
                print(f"  → ... and {len(self.validation_results['award_validation']['issues']) - 5} more")
            all_passed = False
        
        print("="*80)
        if all_passed:
            print("[OK] ALL VALIDATIONS PASSED")
        else:
            print("[X] SOME VALIDATIONS FAILED - Check highlighted cells and comments in output file")
        print("="*80)
    
    def validate(self):
        """Run all validations."""
        print("="*80)
        print("EXCEL VALIDATION TOOL - Quote Win Files")
        print("="*80)
        print(f"Input: {self.input_path}")
        print(f"Output: {self.output_path}")
        print(f"Template: Embedded (Pattern-based validation)")
        print("="*80)
        
        if not self.load_file():
            return False
        
        # Run all validations
        self.validate_headers()
        self.validate_project_name()
        self.validate_filters()
        self.validate_awards()
        
        # Save results
        self.save_results()
        
        # Print summary
        self.print_summary()
        
        return True


def main():
    """Main function to run the validation tool."""
    if len(sys.argv) < 2:
        print("="*80)
        print("Excel Validation Tool - Quote Win Files")
        print("="*80)
        print("\nUsage: python excel_validation_tool.py <input_file> [output_file]")
        print("\nExamples:")
        print("  python excel_validation_tool.py input.xlsx")
        print("  python excel_validation_tool.py input.xlsx output_validated.xlsx")
        print("\nNote: Template is embedded in the tool - no need to provide template file")
        print("="*80)
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    validator = ExcelValidator(input_path, output_path)
    validator.validate()


if __name__ == "__main__":
    main()